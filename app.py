from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import datetime
import os
from dateutil.relativedelta import relativedelta # Para cálculo de datas recorrentes
import calendar # Para obter o número de dias no mês
import json
import google.generativeai as genai
import random
import string
import smtplib
import ssl
from email.mime.text import MIMEText
from flask_migrate import Migrate # Importar Flask-Migrate
import io # Para lidar com arquivos em memória
from openpyxl import Workbook # Para exportar para Excel
from openpyxl.styles import Font, PatternFill # Para estilos no Excel
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE # Para formato de moeda no Excel
from fpdf import FPDF # Para exportar para PDF

# Importa a nova função de gerenciamento de recorrências
# A importação agora é apenas da função, não dos modelos
from manage_recurring import process_subscriptions_and_generate_transactions

app = Flask(__name__)

# --- CONFIGURAÇÃO DA API KEY GEMINI ---
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', 'SUA_CHAVE_DE_API_GEMINI_AQUI')
genai.configure(api_key=GEMINI_API_KEY)


# Configuração da Chave Secreta para Flask-Login e Flask-WTF
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'uma_chave_secreta_muito_complexa_e_aleatoria')

# Configuração do Banco de Dados PostgreSQL (Render) ou SQLite (Local)
DATABASE_URL = os.getenv('DATABASE_URL')
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL or 'sqlite:///' + os.path.join(os.path.abspath(os.path.dirname(__file__)), 'finance.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy()
migrate = Migrate() # Inicialize Migrate sem app e db ainda


login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# --- Configuração de E-mail para Recuperação de Senha ---
EMAIL_USERNAME = os.getenv('EMAIL_USERNAME')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')
EMAIL_SERVER = os.getenv('EMAIL_SERVER', 'smtp.gmail.com')
EMAIL_PORT = int(os.getenv('EMAIL_PORT', 587))

# --- DEBUG: Imprimir valores das variáveis de ambiente de e-mail ---
print(f"DEBUG APP START: EMAIL_USERNAME: {'SET' if EMAIL_USERNAME else 'NOT SET'}")
print(f"DEBUG APP START: EMAIL_PASSWORD: {'SET' if EMAIL_PASSWORD else 'NOT SET'}")
print(f"DEBUG APP START: EMAIL_SERVER: {EMAIL_SERVER}")
print(f"DEBUG APP START: EMAIL_PORT: {EMAIL_PORT}")
# --- FIM DEBUG ---


# --- Definição dos Modelos do Banco de Dados (ORM) ---

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    profile_picture_url = db.Column(db.String(255), nullable=True, default='https://placehold.co/100x100/aabbcc/ffffff?text=PF')
    
    # Campos para recuperação de senha
    recovery_code = db.Column(db.String(6), nullable=True)
    recovery_code_expires_at = db.Column(db.DateTime, nullable=True)

    transactions = db.relationship('Transaction', backref='user', lazy=True, cascade='all, delete-orphan')
    bills = db.relationship('Bill', backref='user', lazy=True, cascade='all, delete-orphan')
    budgets = db.relationship('Budget', backref='user_budget_owner', lazy=True, cascade='all, delete-orphan')
    goals = db.relationship('Goal', backref='user_goal_owner', lazy=True, cascade='all, delete-orphan')
    categories = db.relationship('Category', backref='owner', lazy=True, cascade='all, delete-orphan')
    accounts = db.relationship('Account', backref='owner', lazy=True, cascade='all, delete-orphan')
    # Novo relacionamento para Assinaturas
    subscriptions = db.relationship('Subscription', backref='user', lazy=True, cascade='all, delete-orphan')


    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f"<User {self.email}>"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(10), nullable=False) # 'income' ou 'expense'
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    transactions = db.relationship('Transaction', backref='category', lazy=True)
    budgets = db.relationship('Budget', backref='category', lazy=True) 
    # Novo relacionamento para Assinaturas
    subscriptions = db.relationship('Subscription', backref='category', lazy=True)

    __table_args__ = (db.UniqueConstraint('name', 'type', 'user_id', name='_user_category_type_uc'),)


    def __repr__(self):
        return f"<Category {self.name} ({self.type})>"

class Account(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    balance = db.Column(db.Float, default=0.0)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

    transactions = db.relationship('Transaction', backref='account', lazy=True)
    bills = db.relationship('Bill', backref='account_bill', lazy=True)
    # Novo relacionamento para Assinaturas
    subscriptions = db.relationship('Subscription', backref='account', lazy=True)

    __table_args__ = (db.UniqueConstraint('name', 'user_id', name='_user_account_uc'),)

    def __repr__(self):
        return f"<Account {self.name} (Balance: {self.balance:.2f})>"


class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.String(10), nullable=False) # Armazenar como StringYYYY-MM-DD
    type = db.Column(db.String(10), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=True)
    account_id = db.Column(db.Integer, db.ForeignKey('account.id'), nullable=True)
    goal_id = db.Column(db.Integer, db.ForeignKey('goal.id'), nullable=True)


    def __repr__(self):
        return f"<Transaction {self.description} - {self.amount}>"

class Bill(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    dueDate = db.Column(db.String(10), nullable=False) # Data de vencimento como string
    status = db.Column(db.String(10), nullable=False) # 'pending', 'paid', 'overdue'
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    is_master_recurring_bill = db.Column(db.Boolean, default=False, nullable=False)
    recurring_parent_id = db.Column(db.Integer, db.ForeignKey('bill.id'), nullable=True)
    recurring_child_number = db.Column(db.Integer, nullable=True)
    
    recurring_frequency = db.Column(db.String(20), nullable=True)
    recurring_start_date = db.Column(db.String(10), nullable=True)
    recurring_next_due_date = db.Column(db.String(10), nullable=True)
    recurring_total_occurrences = db.Column(db.Integer, nullable=True)
    recurring_installments_generated = db.Column(db.Integer, nullable=True, default=0)
    is_active_recurring = db.Column(db.Boolean, default=False, nullable=False)
    type = db.Column(db.String(10), nullable=False, default='expense')
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=True)
    account_id = db.Column(db.Integer, db.ForeignKey('account.id'), nullable=True)
    
    payment_transaction_id = db.Column(db.Integer, db.ForeignKey('transaction.id'), nullable=True)
    payment_transaction = db.relationship('Transaction', foreign_keys=[payment_transaction_id], post_update=True)


    def __repr__(self):
        return f"<Bill {self.description} - {self.dueDate} - {self.status}>"

class Budget(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    budget_amount = db.Column(db.Float, nullable=False)
    month_year = db.Column(db.String(7), nullable=False) # 'YYYY-MM'
    current_spent = db.Column(db.Float, default=0.0, nullable=False)

    __table_args__ = (db.UniqueConstraint('user_id', 'category_id', 'month_year', name='_user_category_month_uc'),)

    def __repr__(self):
        return f"<Budget {self.category.name} for {self.month_year}: {self.budget_amount}>"

class Goal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    target_amount = db.Column(db.Float, nullable=False)
    current_amount = db.Column(db.Float, default=0.0, nullable=False)
    due_date = db.Column(db.String(10), nullable=True) #YYYY-MM-DD
    status = db.Column(db.String(20), default='in_progress', nullable=False) # 'in_progress', 'achieved', 'abandoned'
    transactions = db.relationship('Transaction', backref='goal', lazy=True)

    def __repr__(self):
        return f"<Goal {self.name}: {self.current_amount}/{self.target_amount}>"

# NOVO MODELO: Subscription
class Subscription(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    # monthly, quarterly, semi-annually, annually
    billing_cycle = db.Column(db.String(20), nullable=False) 
    # Dia do mês para cobrança (ex: 15 para todo dia 15)
    due_date_of_month = db.Column(db.Integer, nullable=False) 
    next_due_date = db.Column(db.String(10), nullable=False) # YYYY-MM-DD
    status = db.Column(db.String(20), default='active', nullable=False) # active, inactive, cancelled
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=True)
    account_id = db.Column(db.Integer, db.ForeignKey('account.id'), nullable=True)

    def __repr__(self):
        return f"<Subscription {self.name}: R${self.amount} {self.billing_cycle}>"


# --- Funções de Lógica de Negócios (TODAS DEFINIDAS ANTES DAS ROTAS) ---

TODAY_DATE = datetime.date.today()

def get_current_month_year_str():
    return datetime.date.today().strftime('%Y-%m')

def get_month_start_end_dates(month_year_str):
    year, month = map(int, month_year_str.split('-'))
    start_date = datetime.date(year, month, 1)
    end_date = start_date.replace(day=calendar.monthrange(year, month)[1])
    return start_date, end_date

def add_transaction_db(description, amount, date, type, user_id, category_id=None, account_id=None, goal_id=None):
    amount = float(amount)
    date_obj = datetime.datetime.strptime(date, '%Y-%m-%d').date()

    new_transaction = Transaction(
        description=description,
        amount=amount,
        date=date,
        type=type,
        user_id=user_id,
        category_id=category_id,
        account_id=account_id,
        goal_id=goal_id
    )
    db.session.add(new_transaction)
    db.session.flush()

    if account_id:
        account = Account.query.get(account_id)
        if account:
            if type == 'income':
                account.balance += amount
            else: # expense
                account.balance -= amount

    if type == 'expense' and category_id:
        transaction_month_year = date_obj.strftime('%Y-%m')
        budget = Budget.query.filter_by(
            user_id=user_id,
            category_id=category_id,
            month_year=transaction_month_year
        ).first()
        if budget:
            budget.current_spent += amount
            db.session.add(budget)

    if goal_id:
        goal = Goal.query.get(goal_id)
        if goal and goal.user_id == user_id:
            goal.current_amount += amount
            if goal.current_amount >= goal.target_amount:
                goal.status = 'achieved'
                flash(f'Parabéns! Com esta transação, a meta "{goal.name}" foi atingida!', 'success')
            db.session.add(goal)
    
    db.session.commit()
    return new_transaction

def edit_transaction_db(transaction_id, description, amount, date, type, user_id, category_id=None, account_id=None, goal_id=None):
    transaction = Transaction.query.filter_by(id=transaction_id, user_id=user_id).first()
    if not transaction:
        return False

    old_amount = transaction.amount
    old_type = transaction.type
    old_category_id = transaction.category_id
    old_account_id = transaction.account_id
    old_goal_id = transaction.goal_id
    old_date_obj = datetime.datetime.strptime(transaction.date, '%Y-%m-%d').date()

    new_amount = float(amount)
    new_date_obj = datetime.datetime.strptime(date, '%Y-%m-%d').date()

    # Reverte valores antigos
    if old_account_id:
        old_account = Account.query.get(old_account_id)
        if old_account:
            if old_type == 'income': old_account.balance -= old_amount
            else: old_account.balance += old_amount
    if old_type == 'expense' and old_category_id:
        old_budget = Budget.query.filter_by(user_id=user_id, category_id=old_category_id, month_year=old_date_obj.strftime('%Y-%m')).first()
        if old_budget: old_budget.current_spent -= old_amount
    if old_goal_id:
        old_goal = Goal.query.get(old_goal_id)
        if old_goal:
            old_goal.current_amount -= old_amount
            if old_goal.status == 'achieved' and old_goal.current_amount < old_goal.target_amount:
                old_goal.status = 'in_progress'

    # Atualiza a transação
    transaction.description = description
    transaction.amount = new_amount
    transaction.date = date
    transaction.type = type
    transaction.category_id = category_id
    transaction.account_id = account_id
    transaction.goal_id = goal_id

    # Aplica novos valores
    if account_id:
        new_account = Account.query.get(account_id)
        if new_account:
            if type == 'income': new_account.balance += new_amount
            else: new_account.balance -= new_amount
    if type == 'expense' and category_id:
        new_budget = Budget.query.filter_by(user_id=user_id, category_id=category_id, month_year=new_date_obj.strftime('%Y-%m')).first()
        if new_budget: new_budget.current_spent += new_amount
    if goal_id:
        new_goal = Goal.query.get(goal_id)
        if new_goal:
            new_goal.current_amount += new_amount
            if new_goal.current_amount >= new_goal.target_amount:
                new_goal.status = 'achieved'
                flash(f'Parabéns! Com esta transação, a meta "{new_goal.name}" foi atingida!', 'success')

    db.session.commit()
    return True

def delete_transaction_db(transaction_id, user_id):
    transaction = Transaction.query.filter_by(id=transaction_id, user_id=user_id).first()
    if not transaction:
        return False

    if transaction.account_id:
        account = Account.query.get(transaction.account_id)
        if account:
            if transaction.type == 'income':
                account.balance -= transaction.amount
            else: # expense
                account.balance += transaction.amount

    if transaction.type == 'expense' and transaction.category_id:
        transaction_month_year = datetime.datetime.strptime(transaction.date, '%Y-%m-%d').strftime('%Y-%m')
        budget = Budget.query.filter_by(
            user_id=user_id,
            category_id=transaction.category_id,
            month_year=transaction_month_year
        ).first()
        if budget:
            budget.current_spent -= transaction.amount
            db.session.add(budget)

    if transaction.goal_id:
        goal = Goal.query.get(transaction.goal_id)
        if goal and goal.user_id == user_id:
            goal.current_amount -= transaction.amount
            if goal.status == 'achieved' and goal.current_amount < goal.target_amount:
                goal.status = 'in_progress'
            db.session.add(goal)
    
    db.session.delete(transaction)
    db.session.commit()
    return True

def _generate_future_recurring_bills(master_bill):
    print(f"DEBUG: _generate_future_recurring_bills chamada para master_bill ID: {master_bill.id}, Desc: {master_bill.description}")
    
    if not master_bill.id:
        print("ERROR: Master Bill does not have an ID yet. Cannot generate children.")
        return

    generated_count_for_master = 0
    
    Bill.query.filter_by(recurring_parent_id=master_bill.id, user_id=master_bill.user_id, status='pending').delete()
    db.session.commit()

    current_occurrence_date_from_master_start = datetime.datetime.strptime(master_bill.recurring_start_date, '%Y-%m-%d').date()
    
    total_to_generate = master_bill.recurring_total_occurrences if master_bill.recurring_total_occurrences and master_bill.recurring_total_occurrences > 0 else 12
    
    print(f"DEBUG: Total de ocorrências para gerar para {master_bill.description}: {total_to_generate}")

    for i in range(1, total_to_generate + 1):
        occurrence_date_for_child = datetime.datetime.strptime(master_bill.recurring_start_date, '%Y-%m-%d').date()
        
        if master_bill.recurring_frequency == 'monthly' or master_bill.recurring_frequency == 'installments':
            occurrence_date_for_child += relativedelta(months=i-1)
        elif master_bill.recurring_frequency == 'weekly':
            occurrence_date_for_child += relativedelta(weeks=i-1)
        elif master_bill.recurring_frequency == 'yearly':
            occurrence_date_for_child += relativedelta(years=i-1)

        existing_child_item = Bill.query.filter_by(
            recurring_parent_id=master_bill.id,
            dueDate=occurrence_date_for_child.isoformat(),
            recurring_child_number=i,
            user_id=master_bill.user_id
        ).first()

        if not existing_child_item:
            if master_bill.recurring_frequency == 'installments' and master_bill.recurring_total_occurrences > 0:
                child_description = f"{master_bill.description.replace(' (Mestra)', '')} (Parcela {i}/{master_bill.recurring_total_occurrences})"
            else:
                child_description = master_bill.description.replace(' (Mestra)', '')

            new_child_bill_status = 'pending'
            if occurrence_date_for_child < TODAY_DATE:
                new_child_bill_status = 'overdue'

            new_child_bill = Bill(
                description=child_description,
                amount=master_bill.amount,
                dueDate=occurrence_date_for_child.isoformat(),
                status=new_child_bill_status,
                user_id=master_bill.user_id,
                recurring_parent_id=master_bill.id,
                recurring_child_number=i,
                is_master_recurring_bill=False,
                recurring_frequency=None,
                recurring_start_date=None,
                recurring_next_due_date=None,
                recurring_total_occurrences=0,
                recurring_installments_generated=0,
                is_active_recurring=False,
                type=master_bill.type,
                category_id=master_bill.category_id,
                account_id=master_bill.account_id
            )
            db.session.add(new_child_bill)
            generated_count_for_master += 1
            print(f"    Gerada Bill filha: {child_description} em {occurrence_date_for_child.isoformat()} com status {new_child_bill_status}")
        else:
            print(f"    Bill filha já existe para {master_bill.description}, ocorrência {i} em {occurrence_date_for_child.isoformat()}, pulando.")

    master_bill.recurring_installments_generated = generated_count_for_master
    
    # --- CORREÇÃO DA LÓGICA DE recurring_next_due_date ---
    if master_bill.recurring_total_occurrences > 0:
        # Se houver um número fixo de ocorrências, o master bill se desativa após todas serem geradas.
        # O next_due_date deve apontar para a próxima que deveria ser gerada.
        next_occurrence_number = master_bill.recurring_installments_generated + 1
        
        if next_occurrence_number > master_bill.recurring_total_occurrences:
            master_bill.is_active_recurring = False
            master_bill.recurring_next_due_date = None # Não há mais datas futuras
        else:
            next_due_date_for_master = datetime.datetime.strptime(master_bill.recurring_start_date, '%Y-%m-%d').date()
            if master_bill.recurring_frequency == 'monthly' or master_bill.recurring_frequency == 'installments':
                next_due_date_for_master += relativedelta(months=next_occurrence_number - 1)
            elif master_bill.recurring_frequency == 'weekly':
                next_due_date_for_master += relativedelta(weeks=next_occurrence_number - 1)
            elif master_bill.recurring_frequency == 'yearly':
                next_due_date_for_master += relativedelta(years=next_occurrence_number - 1)
            master_bill.recurring_next_due_date = next_due_date_for_master.isoformat()
            print(f"DEBUG: Próximo vencimento da semente '{master_bill.description}' atualizado para: {master_bill.recurring_next_due_date}")
    else: # Indefinido (recurring_total_occurrences é 0)
        # O próximo vencimento é simplesmente o próximo período a partir de HOJE
        next_due_date_for_master = TODAY_DATE
        if master_bill.recurring_frequency == 'monthly' or master_bill.recurring_frequency == 'installments':
            next_due_date_for_master += relativedelta(months=1)
        elif master_bill.recurring_frequency == 'weekly':
            next_due_date_for_master += relativedelta(weeks=1)
        elif master_bill.recurring_frequency == 'yearly':
            next_due_date_for_master += relativedelta(years=1)
        
        master_bill.recurring_next_due_date = next_due_date_for_master.isoformat()
        print(f"DEBUG: Próximo vencimento da semente '{master_bill.description}' (indefinida) atualizado para: {master_bill.recurring_next_due_date}")

    db.session.add(master_bill)
    db.session.commit()
    # --- FIM DA CORREÇÃO ---

def process_recurring_items_on_access(user_id):
    """Processa Bills mestras recorrentes e Assinaturas que precisam gerar novas ocorrências."""
    # Processar Bills Recorrentes
    recurring_seed_bills_to_process = Bill.query.filter(
        Bill.user_id == user_id,
        Bill.is_master_recurring_bill == True,
        Bill.is_active_recurring == True,
        db.cast(Bill.recurring_next_due_date, db.Date) <= TODAY_DATE
    ).all()
    
    print(f"\n--- process_recurring_items_on_access chamada. Processando {len(recurring_seed_bills_to_process)} Bills mestras recorrentes devidas ---")

    for bill_seed in recurring_seed_bills_to_process:
        print(f"    Acionando geração em massa para mestra '{bill_seed.description}' (ID: {bill_seed.id}) por estar vencida.")
        _generate_future_recurring_bills(bill_seed)
    
    # Processar Assinaturas (chamando a função do novo módulo e passando as dependências)
    process_subscriptions_and_generate_transactions(user_id, db, Transaction, Subscription, Account, Category, TODAY_DATE)


def add_bill_db(description, amount, due_date, user_id, 
                is_recurring=False, recurring_frequency=None, recurring_total_occurrences=0, bill_type='expense', category_id=None, account_id=None):
    
    amount = float(amount)
    if is_recurring and recurring_frequency == 'installments' and (recurring_total_occurrences is None or recurring_total_occurrences < 1):
        recurring_total_occurrences = 1 
    elif not is_recurring or recurring_frequency != 'installments':
        recurring_total_occurrences = 0 

    new_bill = Bill(
        description=description,
        amount=amount,
        dueDate=due_date,
        status='pending',
        user_id=user_id,
        is_master_recurring_bill=is_recurring,
        recurring_frequency=recurring_frequency if is_recurring else None,
        recurring_start_date=due_date if is_recurring else None,
        recurring_next_due_date=due_date if is_recurring else None,
        recurring_total_occurrences=recurring_total_occurrences,
        recurring_installments_generated=0,
        is_active_recurring=is_recurring,
        type=bill_type,
        category_id=category_id,
        account_id=account_id
    )
    db.session.add(new_bill)
    db.session.commit()
    
    if new_bill.is_master_recurring_bill and new_bill.is_active_recurring:
        _generate_future_recurring_bills(new_bill)

def pay_bill_db(bill_id, user_id, payment_account_id): # Adicionado payment_account_id
    bill = Bill.query.filter_by(id=bill_id, user_id=user_id).first()
    if not bill:
        return False

    if bill.status == 'paid':
        print(f"DEBUG: Bill {bill_id} já está paga.")
        return False

    account = Account.query.get(payment_account_id) # Usa payment_account_id
    if not account or account.user_id != user_id:
        print(f"ERROR: Conta {payment_account_id} não encontrada ou não pertence ao usuário {user_id}.")
        flash(f'Saldo insuficiente na conta {account.name} para pagar a conta "{bill.description}".', 'danger')
        return False

    if account.balance < bill.amount:
        print(f"ERROR: Saldo insuficiente na conta {account.name} para a conta {bill.description}.")
        flash(f'Saldo insuficiente na conta {account.name} para pagar a conta "{bill.description}".', 'danger')
        return False

    category_for_payment_id = bill.category_id
    if not category_for_payment_id:
        default_cat = Category.query.filter_by(user_id=user_id, name='Contas Fixas', type='expense').first()
        if not default_cat:
            default_cat = Category.query.filter_by(user_id=user_id, name='Outras Despesas', type='expense').first()
        if default_cat:
            category_for_payment_id = default_cat.id
        else:
            print("AVISO: Nenhuma categoria de despesa padrão encontrada para o pagamento da conta.")
            flash("Aviso: Categoria padrão para pagamento de conta não encontrada. A transação pode não ser categorizada corretamente.", 'warning')
            category_for_payment_id = None # Garante que seja None se não encontrar
            

    new_payment_transaction = add_transaction_db(
        description=f"Pagamento: {bill.description}",
        amount=bill.amount,
        date=TODAY_DATE.isoformat(),
        type='expense',
        user_id=user_id,
        category_id=category_for_payment_id,
        account_id=payment_account_id # Usa payment_account_id
    )

    if new_payment_transaction:
        bill.payment_transaction_id = new_payment_transaction.id
        bill.status = 'paid'
        db.session.add(bill)
        
        master_bill_to_process = None
        if bill.is_master_recurring_bill:
            master_bill_to_process = bill
        elif bill.recurring_parent_id:
            master_bill_to_process = Bill.query.filter_by(id=bill.recurring_parent_id, user_id=user_id, is_master_recurring_bill=True).first()

        if master_bill_to_process and master_bill_to_process.is_active_recurring:
            _generate_future_recurring_bills(master_bill_to_process)

        db.session.commit()
        return True
    
    return False


def reschedule_bill_db(bill_id, new_date, user_id):
    bill = Bill.query.filter_by(id=bill_id, user_id=user_id).first()
    if bill:
        bill.dueDate = new_date
        if datetime.datetime.strptime(new_date, '%Y-%m-%d').date() >= TODAY_DATE and bill.status == 'overdue':
            bill.status = 'pending'
        db.session.add(bill)
        db.session.commit()
        return True
    return False

def delete_bill_db(bill_id, user_id):
    bill = Bill.query.filter_by(id=bill_id, user_id=user_id).first()
    if not bill:
        return False

    print(f"DEBUG: Deleting Bill ID: {bill.id}, Desc: '{bill.description}', Is Master: {bill.is_master_recurring_bill}, Parent ID: {bill.recurring_parent_id}")

    if bill.status == 'paid' and bill.payment_transaction_id:
        payment_transaction = Transaction.query.get(bill.payment_transaction_id)
        if payment_transaction and payment_transaction.user_id == user_id:
            if payment_transaction.account_id:
                account = Account.query.get(payment_transaction.account_id)
                if account:
                    account.balance += payment_transaction.amount
                    db.session.add(account)

            if payment_transaction.type == 'expense' and payment_transaction.category_id:
                transaction_month_year = datetime.datetime.strptime(payment_transaction.date, '%Y-%m-%d').strftime('%Y-%m')
                budget = Budget.query.filter_by(
                    user_id=user_id,
                    category_id=payment_transaction.category_id,
                    month_year=transaction_month_year
                ).first()
                if budget:
                    budget.current_spent -= payment_transaction.amount
                    db.session.add(budget)
                    print(f"DEBUG: Budget {budget.category.name} reverted for deleted payment. New spent: {budget.current_spent}")
            db.session.delete(payment_transaction)
            print(f"DEBUG: Deleted associated payment transaction ID: {payment_transaction.id}")
        else:
            print(f"WARNING: Associated payment transaction ID {bill.payment_transaction_id} not found or doesn't belong to user, cannot revert.")

    if bill.is_master_recurring_bill:
        child_bills = Bill.query.filter_by(recurring_parent_id=bill.id, user_id=user_id).all()
        for child_bill in child_bills:
            db.session.delete(child_bill)
            print(f"DEBUG: Deleting child bill: ID {child_bill.id}, Desc: '{child_bill.description}'")
        print(f"DEBUG: Deleted {len(child_bills)} child bills for master '{bill.description}'.")
        
        db.session.delete(bill)
        db.session.commit()
        print(f"DEBUG: Master bill '{bill.description}' (ID: {bill.id}) and its children deleted.")
        return True

    elif bill.recurring_parent_id:
        master_bill = Bill.query.filter_by(id=bill.recurring_parent_id, user_id=user_id, is_master_recurring_bill=True).first()
        if master_bill:
            print(f"DEBUG: Child bill '{bill.description}' (ID: {bill.id}) being deleted. Attempting to cancel master series '{master_bill.description}'.")
            
            master_bill.is_active_recurring = False
            master_bill.recurring_frequency = None
            master_bill.recurring_start_date = None
            master_bill.recurring_next_due_date = None
            master_bill.recurring_total_occurrences = 0
            master_bill.recurring_installments_generated = 0
            db.session.add(master_bill)

            all_children_of_master = Bill.query.filter_by(recurring_parent_id=master_bill.id, user_id=user_id).all()
            for child in all_children_of_master:
                if child.status == 'paid' and child.payment_transaction_id:
                    child.payment_transaction_id = None
                    db.session.add(child)
                db.session.delete(child)
                print(f"DEBUG: Deleting another child bill (from master): ID {child.id}, Desc: '{child.description}'")
            
            db.session.commit()
            print(f"DEBUG: Recurring series for master '{master_bill.description}' (ID: {master_bill.id}) cancelled and all its children deleted.")
            return True
        else:
            print(f"DEBUG: Child bill '{bill.description}' (ID: {bill.id}) deleted, but master recurring bill (ID: {bill.recurring_parent_id}) not found or is not a master.")
            db.session.delete(bill)
            db.session.commit()
            return True
    else:
        db.session.delete(bill)
        db.session.commit()
        print(f"DEBUG: Non-recurring bill '{bill.description}' (ID: {bill.id}) deleted.")
        return True


def edit_bill_db(bill_id, description, amount, dueDate, user_id, 
                   is_recurring=False, recurring_frequency=None, recurring_total_occurrences=0, is_active_recurring=False, bill_type='expense', category_id=None, account_id=None):
    bill = Bill.query.filter_by(id=bill_id, user_id=user_id).first()
    if not bill:
        return False

    bill.description = description
    bill.amount = float(amount)
    bill.dueDate = dueDate
    bill.type = bill_type
    bill.category_id = category_id
    bill.account_id = account_id

    if bill.status == 'paid':
        print(f"WARNING: Bill {bill.id} is already paid. Changes to amount/type/category won't update past transaction/budget.")

    if bill.is_master_recurring_bill or is_recurring:
        old_is_master = bill.is_master_recurring_bill
        old_is_active = bill.is_active_recurring
        old_frequency = bill.recurring_frequency
        old_total_occurrences = bill.recurring_total_occurrences

        bill.is_master_recurring_bill = is_recurring
        bill.is_active_recurring = is_active_recurring
        bill.recurring_frequency = recurring_frequency if is_recurring else None
        bill.recurring_total_occurrences = recurring_total_occurrences if is_recurring and recurring_frequency == 'installments' else 0
        
        regenerate = False
        if is_recurring:
            if not old_is_master:
                regenerate = True
            elif is_active_recurring and (not old_is_active or old_frequency != recurring_frequency or old_total_occurrences != recurring_total_occurrences):
                regenerate = True
        
        if regenerate:
            print(f"DEBUG: Editing master Bill '{bill.description}'. Recurring parameters changed or re-activated. Regenerating future occurrences.")
            _generate_future_recurring_bills(bill)
        elif old_is_master and not is_recurring:
            bill.is_active_recurring = False
            bill.recurring_frequency = None
            bill.recurring_next_due_date = None
            bill.recurring_total_occurrences = 0
            bill.recurring_installments_generated = 0
            Bill.query.filter_by(recurring_parent_id=bill.id, user_id=user_id, status='pending').delete()
            print(f"DEBUG: Master Bill '{bill.description}' deactivated, future child bills deleted.")
        elif old_is_active and not is_active_recurring and is_recurring:
                print(f"DEBUG: Master Bill '{bill.description}' manually set to inactive.")

    else:
        bill.is_master_recurring_bill = False
        bill.is_active_recurring = False
        bill.recurring_frequency = None
        bill.recurring_start_date = None
        bill.recurring_next_due_date = None
        bill.recurring_total_occurrences = 0
        bill.recurring_installments_generated = 0

    db.session.commit()
    return True

def get_dashboard_data_db(user_id):
    current_month_start = TODAY_DATE.replace(day=1)
    next_month_start = (current_month_start + relativedelta(months=1))
    
    start_date_str = current_month_start.isoformat()
    end_date_str = next_month_start.isoformat()

    total_balance = db.session.query(db.func.sum(Account.balance)).filter_by(user_id=user_id).scalar() or 0.0

    monthly_income = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.user_id == user_id,
        Transaction.type == 'income',
        Transaction.date >= start_date_str,
        Transaction.date < end_date_str
    ).scalar() or 0.0

    monthly_expenses = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.user_id == user_id,
        Transaction.type == 'expense',
        Transaction.date >= start_date_str,
        Transaction.date < end_date_str
    ).scalar() or 0.0
    
    monthly_pending_bills_amount = db.session.query(db.func.sum(Bill.amount)).filter(
        Bill.user_id == user_id,
        Bill.status == 'pending',
        Bill.is_master_recurring_bill == False,
        db.cast(Bill.dueDate, db.Date) < next_month_start,
        db.cast(Bill.dueDate, db.Date) >= current_month_start.replace(day=1)
    ).scalar() or 0.0

    all_pending_bills_list = Bill.query.filter(
        Bill.user_id == user_id,
        Bill.status == 'pending',
        Bill.is_master_recurring_bill == False,
        db.cast(Bill.dueDate, db.Date) < next_month_start
    ).order_by(db.cast(Bill.dueDate, db.Date).asc()).all()


    return {
        'balance': total_balance,
        'totalIncome': monthly_income,
        'totalExpenses': monthly_expenses,
        'totalPendingBills': monthly_pending_bills_amount,
        'pendingBillsList': all_pending_bills_list
    }

# --- FUNÇÕES GEMINI ---
def generate_text_with_gemini(prompt_text):
    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content(prompt_text)
        return response.text
    except Exception as e:
        print(f"ERROR: Erro ao chamar Gemini API: {e}")
        return "Não foi possível gerar uma sugestão/resumo no momento. Verifique sua chave de API e conexão."

# --- NOVAS FUNÇÕES PARA ORÇAMENTOS E METAS (DB operations) ---
def add_budget_db(user_id, category_id, budget_amount, month_year):
    existing_budget = Budget.query.filter_by(
        user_id=user_id,
        category_id=category_id,
        month_year=month_year
    ).first()

    if existing_budget:
        existing_budget.budget_amount = float(budget_amount)
    else:
        new_budget = Budget(
            user_id=user_id,
            category_id=category_id,
            budget_amount=float(budget_amount),
            month_year=month_year
        )
        db.session.add(new_budget)
    
    start_date, end_date = get_month_start_end_dates(month_year)
    start_date_str = start_date.isoformat()
    end_date_str = end_date.isoformat()

    total_spent_in_category = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.user_id == user_id,
        Transaction.category_id == category_id,
        Transaction.type == 'expense',
        Transaction.date >= start_date_str,
        Transaction.date <= end_date_str
    ).scalar() or 0.0

    if existing_budget:
        existing_budget.current_spent = total_spent_in_category
        db.session.add(existing_budget)
    else: # new_budget
        new_budget.current_spent = total_spent_in_category
        db.session.add(new_budget)

    db.session.commit()
    return True

def edit_budget_db(budget_id, user_id, budget_amount=None):
    budget = Budget.query.filter_by(id=budget_id, user_id=user_id).first()
    if budget:
        if budget_amount is not None:
            budget.budget_amount = float(budget_amount)
        
        start_date, end_date = get_month_start_end_dates(budget.month_year)
        start_date_str = start_date.isoformat()
        end_date_str = end_date.isoformat()
        
        total_spent_in_category = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.user_id == user_id,
            Transaction.category_id == budget.category_id,
            Transaction.type == 'expense',
            Transaction.date >= start_date_str,
            Transaction.date <= end_date_str
        ).scalar() or 0.0
        budget.current_spent = total_spent_in_category
        db.session.commit()
        return True
    return False

def delete_budget_db(budget_id, user_id):
    budget = Budget.query.filter_by(id=budget_id, user_id=user_id).first()
    if budget:
        db.session.delete(budget)
        db.session.commit()
        return True
    return False

def add_goal_db(user_id, name, target_amount, due_date=None):
    new_goal = Goal(
        user_id=user_id,
        name=name,
        target_amount=float(target_amount),
        due_date=due_date if due_date else None,
        status='in_progress'
    )
    db.session.add(new_goal)
    db.session.commit()
    return True

def edit_goal_db(goal_id, user_id, name=None, target_amount=None, current_amount=None, due_date=None, status=None):
    goal = Goal.query.filter_by(id=goal_id, user_id=user_id).first()
    if goal:
        if name:
            goal.name = name
        if target_amount is not None:
            goal.target_amount = float(target_amount)
        if current_amount is not None:
            goal.current_amount = float(current_amount)
        if due_date:
            goal.due_date = due_date
        if status:
            goal.status = status
        db.session.commit()
        return True
    return False

def delete_goal_db(goal_id, user_id):
    goal = Goal.query.filter_by(id=goal_id, user_id=user_id).first()
    if goal:
        db.session.delete(goal)
        db.session.commit()
        return True
    return False

def contribute_to_goal_db(goal_id, user_id, amount, source_account_id): # Adicionado source_account_id
    goal = Goal.query.filter_by(id=goal_id, user_id=user_id).first()
    if not goal:
        return False
    
    amount_to_add = float(amount)
    if amount_to_add <= 0:
        flash('O valor da contribuição deve ser maior que zero.', 'danger')
        return False

    source_account = Account.query.get(source_account_id) # Usa source_account_id
    if not source_account or source_account.user_id != user_id:
        print(f"ERROR: Conta de origem {source_account_id} não encontrada ou não pertence ao usuário {user_id}.")
        flash('Conta de origem inválida ou não pertence a você.', 'danger')
        return False

    if source_account.balance < amount_to_add:
        print(f"ERROR: Saldo insuficiente na conta {source_account.name} para contribuir com a meta {goal.name}.")
        flash(f'Saldo insuficiente na conta {source_account.name} para pagar a conta "{goal.name}".', 'danger')
        return False

    if goal.current_amount + amount_to_add >= goal.target_amount:
        amount_to_add_actual = goal.target_amount - goal.current_amount
        goal.status = 'achieved'
        flash(f'Parabéns! Com esta contribuição, a meta "{goal.name}" foi atingida!', 'success')
    else:
        amount_to_add_actual = amount_to_add
        flash(f'Contribuição de R$ {amount_to_add_actual:.2f} adicionada à meta "{goal.name}".', 'success')
        
    goal.current_amount += amount_to_add_actual
    db.session.add(goal)

    poupanca_metas_category = Category.query.filter_by(name='Poupança para Metas', type='expense', user_id=user_id).first()
    
    if poupanca_metas_category:
        new_transaction = Transaction(
            description=f"Contribuição para Meta: {goal.name}",
            amount=amount_to_add_actual,
            date=TODAY_DATE.isoformat(),
            type='expense',
            user_id=user_id,
            category_id=poupanca_metas_category.id,
            account_id=source_account_id # Usa source_account_id
        )
        db.session.add(new_transaction)
        source_account.balance -= amount_to_add_actual # Debita da conta de origem
        db.session.add(source_account)

        transaction_month_year = TODAY_DATE.strftime('%Y-%m')
        budget = Budget.query.filter_by(
            user_id=user_id,
            category_id=poupanca_metas_category.id,
            month_year=transaction_month_year
        ).first()
        if budget:
            budget.current_spent += amount_to_add_actual
            db.session.add(budget)
            print(f"DEBUG: Budget for category '{poupanca_metas_category.name}' updated with goal contribution. New spent: {budget.current_spent}")
        else:
            print(f"DEBUG: No budget found for category 'Poupança para Metas' for {transaction_month_year} to update.")
    else:
        print("WARNING: Could not create transaction for goal contribution (missing category 'Poupança para Metas').")
        flash("Aviso: Categoria 'Poupança para Metas' não encontrada. A transação da meta não foi registrada.", 'warning')

    db.session.commit()
    return True


# --- Funções de Gerenciamento de Contas ---
def add_account_db(user_id, name, initial_balance):
    """Adiciona uma nova conta para o usuário."""
    existing_account = Account.query.filter_by(user_id=user_id, name=name).first()
    if existing_account:
        return False # Conta com o mesmo nome já existe para este usuário
    
    new_account = Account(
        user_id=user_id,
        name=name,
        balance=float(initial_balance)
    )
    db.session.add(new_account)
    db.session.commit()
    return True

def edit_account_db(account_id, user_id, new_name=None, new_balance=None):
    """Edita uma conta existente do usuário."""
    account = Account.query.filter_by(id=account_id, user_id=user_id).first()
    if not account:
        return False
    
    if new_name and new_name != account.name:
        # Verifica se o novo nome já existe para evitar duplicatas
        existing_account_with_new_name = Account.query.filter_by(user_id=user_id, name=new_name).first()
        if existing_account_with_new_name and existing_account_with_new_name.id != account_id:
            return False # Já existe outra conta com este nome
        account.name = new_name
    
    if new_balance is not None:
        account.balance = float(new_balance)
    
    db.session.commit()
    return True

def delete_account_db(account_id, user_id):
    """Exclui uma conta e desvincula suas transações."""
    account = Account.query.filter_by(id=account_id, user_id=user_id).first()
    if not account:
        return False
    
    # Antes de deletar a conta, desvincular todas as transações e contas a pagar associadas
    # Definindo account_id para None
    Transaction.query.filter_by(account_id=account_id, user_id=user_id).update({'account_id': None})
    Bill.query.filter_by(account_id=account_id, user_id=user_id).update({'account_id': None})
    Subscription.query.filter_by(account_id=account_id, user_id=user_id).update({'account_id': None}) # NOVO: Desvincular assinaturas
    
    db.session.delete(account)
    db.session.commit()
    return True

def transfer_funds_db(user_id, source_account_id, destination_account_id, amount):
    """Transfere fundos entre duas contas do usuário."""
    amount = float(amount)
    if amount <= 0:
        flash('O valor da transferência deve ser maior que zero.', 'danger')
        return False

    if source_account_id == destination_account_id:
        flash('A conta de origem e a conta de destino não podem ser a mesma.', 'danger')
        return False

    source_account = Account.query.filter_by(id=source_account_id, user_id=user_id).first()
    destination_account = Account.query.filter_by(id=destination_account_id, user_id=user_id).first()

    if not source_account or not destination_account:
        flash('Conta de origem ou destino não encontrada.', 'danger')
        return False

    if source_account.balance < amount:
        flash(f'Saldo insuficiente na conta de origem ({source_account.name}) para realizar a transferência.', 'danger')
        return False

    # Debitar da conta de origem
    source_account.balance -= amount
    db.session.add(source_account)

    # Creditar na conta de destino
    destination_account.balance += amount
    db.session.add(destination_account)

    # Criar transação de despesa para a conta de origem
    # Usamos category_id=None para transferências, pois não são despesas reais, apenas movimentação de fundos.
    add_transaction_db(
        description=f"Transferência para {destination_account.name}",
        amount=amount,
        date=TODAY_DATE.isoformat(),
        type='expense',
        user_id=user_id,
        category_id=None, # Transferências não afetam categorias de despesa/receita
        account_id=source_account_id
    )

    # Criar transação de receita para a conta de destino
    add_transaction_db(
        description=f"Transferência de {source_account.name}",
        amount=amount,
        date=TODAY_DATE.isoformat(),
        type='income',
        user_id=user_id,
        category_id=None, # Transferências não afetam categorias de despesa/receita
        account_id=destination_account_id
    )

    db.session.commit()
    flash('Transferência realizada com sucesso!', 'success')
    return True

# --- Funções para Relatórios Detalhados ---
def get_detailed_report_data_db(user_id, start_date_str, end_date_str, transaction_type=None, category_id=None):
    """
    Busca dados detalhados para relatórios com base em filtros.
    Retorna dados para gráficos e um resumo para a IA.
    """
    
    # Validação de datas
    try:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return {'error': 'Formato de data inválido.'}

    # Query base para transações
    transactions_query = Transaction.query.filter(
        Transaction.user_id == user_id,
        Transaction.date >= start_date_str,
        Transaction.date <= end_date_str # Usar <= para incluir o dia final
    )

    if transaction_type and transaction_type in ['income', 'expense']:
        transactions_query = transactions_query.filter(Transaction.type == transaction_type)
    
    if category_id:
        transactions_query = transactions_query.filter(Transaction.category_id == category_id)

    transactions = transactions_query.all()

    # 1. Dados para Gráfico de Despesas por Categoria (Pie Chart)
    expenses_by_category = {}
    total_expenses_in_period = 0.0
    for t in transactions:
        if t.type == 'expense' and t.category:
            category_name = t.category.name
            expenses_by_category[category_name] = expenses_by_category.get(category_name, 0) + t.amount
            total_expenses_in_period += t.amount
    
    expenses_chart_data = {
        'labels': list(expenses_by_category.keys()),
        'values': list(expenses_by_category.values())
    }

    # 2. Dados para Gráfico de Evolução do Patrimônio Líquido (Line Chart)
    # Isso é mais complexo, pois exige o saldo das contas ao longo do tempo.
    # Uma abordagem simplificada é calcular o saldo líquido cumulativo a partir de um ponto inicial.
    # Para um relatório detalhado, podemos pegar o saldo inicial de todas as contas e aplicar as transações.
    
    # Pega o saldo inicial de todas as contas no início do período
    initial_balance_sum = db.session.query(db.func.sum(Account.balance)).filter(
        Account.user_id == user_id
    ).scalar() or 0.0 # Saldo atual de todas as contas

    # Para a evolução do patrimônio, precisamos de pontos de dados ao longo do tempo.
    # Vamos gerar um ponto por dia dentro do período selecionado.
    net_worth_labels = []
    net_worth_values = []

    # Começa com o saldo total atual e "reverte" as transações para o passado
    # Ou, uma abordagem mais simples: recalcular o saldo para cada ponto no tempo
    
    # Pegar todas as transações do usuário até a data final do relatório
    all_user_transactions_until_end_date = Transaction.query.filter(
        Transaction.user_id == user_id,
        Transaction.date <= end_date_str
    ).order_by(Transaction.date.asc()).all()

    # Calcula o saldo inicial da primeira transação ou 0 se não houver transações
    current_net_worth = 0.0
    if all_user_transactions_until_end_date:
        # Encontra o primeiro saldo conhecido ou assume 0 antes da primeira transação
        first_transaction_date = datetime.datetime.strptime(all_user_transactions_until_end_date[0].date, '%Y-%m-%d').date()
        
        # Se o relatório começa antes da primeira transação, o saldo inicial é 0.
        # Se o relatório começa depois da primeira transação, precisamos calcular o saldo até o start_date.
        
        # Calcula o saldo até o dia anterior ao start_date do relatório
        balance_before_report_start = 0.0
        for t in all_user_transactions_until_end_date:
            t_date = datetime.datetime.strptime(t.date, '%Y-%m-%d').date()
            if t_date < start_date:
                if t.type == 'income':
                    balance_before_report_start += t.amount
                else:
                    balance_before_report_start -= t.amount
        current_net_worth = balance_before_report_start
    
    # Itera pelos dias/meses dentro do período do relatório
    current_date_iter = start_date
    while current_date_iter <= end_date:
        net_worth_labels.append(current_date_iter.strftime('%d/%m/%Y'))
        
        # Adiciona o efeito das transações do dia atual
        for t in all_user_transactions_until_end_date:
            t_date = datetime.datetime.strptime(t.date, '%Y-%m-%d').date()
            if t_date == current_date_iter:
                if t.type == 'income':
                    current_net_worth += t.amount
                else:
                    current_net_worth -= t.amount
        net_worth_values.append(current_net_worth)
        current_date_iter += datetime.timedelta(days=1) # Incrementa por dia para mais granularidade

    net_worth_chart_data = {
        'labels': net_worth_labels,
        'values': net_worth_values
    }


    # 3. Resumo para IA
    total_income_in_period = sum(t.amount for t in transactions if t.type == 'income')
    total_expenses_in_period_for_ai = sum(t.amount for t in transactions if t.type == 'expense') # Use this for AI
    balance_in_period = total_income_in_period - total_expenses_in_period_for_ai

    ai_summary_data = {
        'start_date': start_date_str,
        'end_date': end_date_str,
        'total_income': total_income_in_period,
        'total_expenses': total_expenses_in_period_for_ai,
        'balance': balance_in_period,
        'expenses_by_category': expenses_by_category, # Detalhes para IA
        'transaction_count': len(transactions)
    }

    return {
        'expenses_by_category_chart': expenses_chart_data,
        'net_worth_evolution_chart': net_worth_chart_data,
        'ai_summary': ai_summary_data
    }


# --- Funções de Email ---
def generate_recovery_code(length=6):
    """Gera um código alfanumérico aleatório."""
    characters = string.ascii_uppercase + string.digits
    return ''.join(random.choice(characters) for i in range(length))

def send_recovery_email(recipient_email, recovery_code):
    """Envia o código de recuperação para o e-mail do usuário."""
    if not EMAIL_USERNAME or not EMAIL_PASSWORD:
        print("ERROR: send_recovery_email - Credenciais de email (EMAIL_USERNAME ou EMAIL_PASSWORD) não configuradas. Verifique as variáveis de ambiente.")
        return False

    sender_email = EMAIL_USERNAME
    sender_password = EMAIL_PASSWORD
    smtp_server = EMAIL_SERVER
    smtp_port = EMAIL_PORT

    message = MIMEText(
        f"Seu código de recuperação de senha é: {recovery_code}\n"
        "Este código é válido por 10 minutos. Se você não solicitou esta redefinição, por favor, ignore este e-mail."
    )
    message["Subject"] = "Código de Recuperação de Senha - Gestão Financeira Pessoal"
    message["From"] = sender_email
    message["To"] = recipient_email
    
    try:
        print(f"DEBUG: send_recovery_email - Tentando conectar a {smtp_server}:{smtp_port} com usuário {sender_email}")
        context = ssl.create_default_context()
        
        # --- CORREÇÃO: Usar smtplib.SMTP com starttls() para a porta 587 ---
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls(context=context) # Inicia a criptografia TLS
            server.login(sender_email, sender_password)
            print(f"DEBUG: send_recovery_email - Login SMTP bem-sucedido para {sender_email}")
            server.sendmail(sender_email, recipient_email, message.as_string())
        # --- FIM CORREÇÃO ---

        print(f"DEBUG: send_recovery_email - Email de recuperação enviado para {recipient_email}")
        return True
    except smtplib.SMTPAuthenticationError as auth_err:
        print(f"ERROR: send_recovery_email - SMTPAuthenticationError: Falha de autenticação. Verifique o EMAIL_USERNAME e a SENHA DE APLICAÇÃO (App Password) no Render. Erro: {auth_err}")
        return False
    except smtplib.SMTPServerDisconnected as disconnect_err:
        print(f"ERROR: send_recovery_email - SMTPServerDisconnected: Servidor desconectado inesperadamente. Verifique EMAIL_SERVER e EMAIL_PORT. Erro: {disconnect_err}")
        return False
    except smtplib.SMTPException as smtp_err:
        print(f"ERROR: send_recovery_email - SMTPException: Erro geral SMTP. Pode ser problema de conexão, firewall ou servidor SMTP. Erro: {smtp_err}")
        return False
    except Exception as e:
        print(f"ERROR: send_recovery_email - Erro inesperado ao enviar email: {type(e).__name__} - {e}")
        return False


# --- ROTAS DA APLICAÇÃO ---

@app.route('/')
@login_required
def index():
    """Rota principal do dashboard."""
    # Chamada atualizada para passar os modelos e a instância do db
    process_recurring_items_on_access(current_user.id) 
    
    dashboard_data = get_dashboard_data_db(current_user.id)
    
    transactions_query_obj = Transaction.query.filter_by(user_id=current_user.id) 

    transaction_type_filter = request.args.get('transaction_type')
    if transaction_type_filter and transaction_type_filter in ['income', 'expense']:
        transactions_query_obj = transactions_query_obj.filter_by(type=transaction_type_filter)

    start_date_filter = request.args.get('start_date')
    end_date_filter = request.args.get('end_date')
    if start_date_filter:
        transactions_query_obj = transactions_query_obj.filter(Transaction.date >= start_date_filter)
    if end_date_filter:
        transactions_query_obj = transactions_query_obj.filter(Transaction.date <= end_date_filter)

    category_filter_id = request.args.get('category_filter', type=int)
    if category_filter_id:
        transactions_query_obj = transactions_query_obj.filter_by(category_id=category_filter_id)

    sort_by_transactions = request.args.get('sort_by_transactions', 'date')
    order_transactions = request.args.get('order_transactions', 'desc')

    if sort_by_transactions == 'date':
        if order_transactions == 'asc':
            transactions_query_obj = transactions_query_obj.order_by(Transaction.date.asc())
        else:
            transactions_query_obj = transactions_query_obj.order_by(Transaction.date.desc())
    elif sort_by_transactions == 'amount':
        if order_transactions == 'asc':
            transactions_query_obj = transactions_query_obj.order_by(Transaction.amount.asc())
        else:
            transactions_query_obj = transactions_query_obj.order_by(Transaction.amount.desc())
            
    all_transactions = transactions_query_obj.all()
    
    income_transactions = [t for t in all_transactions if t.type == 'income']
    expense_transactions = [t for t in all_transactions if t.type == 'expense']


    bills_query_obj = Bill.query.filter( 
        Bill.user_id == current_user.id,
        Bill.is_master_recurring_bill == False
    )
    bill_status_filter = request.args.get('bill_status')
    if bill_status_filter and bill_status_filter in ['pending', 'paid', 'overdue']:
        if bill_status_filter == 'overdue':
            bills_query_obj = bills_query_obj.filter(db.cast(Bill.dueDate, db.Date) < TODAY_DATE, Bill.status == 'pending')
        else:
            bills_query_obj = bills_query_obj.filter_by(status=bill_status_filter)
    elif not bill_status_filter:
            bills_query_obj = bills_query_obj.filter_by(status='pending')
            
    filtered_bills = bills_query_obj.order_by(db.cast(Bill.dueDate, db.Date).asc()).all()

    all_categories_formatted = [(c.id, c.type, c.name) for c in Category.query.filter_by(user_id=current_user.id).all()]
    
    user_accounts = Account.query.filter_by(user_id=current_user.id).all()
    accounts_json = [{'id': acc.id, 'name': acc.name, 'balance': acc.balance} for acc in user_accounts]

    current_month_year = get_current_month_year_str()
    budgets_with_alerts = []
    all_budgets_for_month = Budget.query.filter_by(user_id=current_user.id, month_year=current_month_year).all()

    for budget in all_budgets_for_month:
        if budget.budget_amount > 0:
            percentage_spent = (budget.current_spent / budget.budget_amount * 100)
            
            alert_data = {
                'category_name': budget.category.name,
                'percentage_spent': round(percentage_spent)
            }

            if percentage_spent > 100:
                alert_data['status'] = 'danger'
                alert_data['message'] = f"Alerta: Você ultrapassou em {round(percentage_spent - 100)}% o orçamento para"
                budgets_with_alerts.append(alert_data)
            elif percentage_spent >= 80:
                alert_data['status'] = 'warning'
                alert_data['message'] = f"Atenção: Você já utilizou {round(percentage_spent)}% do seu orçamento para"
                budgets_with_alerts.append(alert_data)

    active_goals = Goal.query.filter_by(user_id=current_user.id, status='in_progress').order_by(Goal.name.asc()).all()
    goals_json = [{'id': goal.id, 'name': goal.name} for goal in active_goals]

    show_new_budget_alert = False
    if not all_budgets_for_month:
        show_new_budget_alert = True


    return render_template(
        'index.html',
        dashboard=dashboard_data,
        bills=filtered_bills,
        income_transactions=income_transactions,
        expense_transactions=expense_transactions,
        current_date=TODAY_DATE.isoformat(),
        current_user=current_user,
        current_transaction_type_filter=transaction_type_filter,
        current_bill_status_filter=bill_status_filter,
        current_sort_by_transactions=sort_by_transactions,
        current_order_transactions=order_transactions,
        current_start_date=start_date_filter,
        current_end_date=end_date_filter,
        all_categories=all_categories_formatted,
        current_category_filter=category_filter_id,
        accounts=user_accounts,
        accounts_json=accounts_json,
        budgets_with_alerts=budgets_with_alerts,
        active_goals=active_goals,
        goals_json=goals_json,
        show_new_budget_alert=show_new_budget_alert
    )

@app.route('/add_transaction', methods=['POST'])
@login_required
def handle_add_transaction():
    description = request.form['description']
    amount = request.form['amount']
    date = request.form['date']
    transaction_type = request.form['type']
    category_id = request.form.get('category_id', type=int)
    account_id = request.form.get('account_id', type=int)
    goal_id = request.form.get('goal_id', type=int) if request.form.get('goal_id') else None

    add_transaction_db(description, amount, date, transaction_type, current_user.id, category_id, account_id, goal_id)
    flash('Transação adicionada com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/add_bill', methods=['POST'])
@login_required
def handle_add_bill():
    description = request.form['bill_description']
    amount = request.form['bill_amount']
    due_date = request.form['bill_due_date']
    
    is_recurring = request.form.get('is_recurring_bill') == 'on'
    recurring_frequency = request.form.get('recurring_frequency_bill')
    recurring_total_occurrences = request.form.get('recurring_total_occurrences_bill', type=int)
    bill_type = request.form['bill_type']
    category_id = request.form.get('bill_category_id', type=int)
    # account_id removido do formulário, mas ainda pode ser passado como None ou padrão se necessário
    account_id = None # Ou um ID de conta padrão se houver um conceito de conta "principal" para contas a pagar

    add_bill_db(description, amount, due_date, current_user.id, 
                is_recurring, recurring_frequency, recurring_total_occurrences, bill_type, category_id, account_id)
    flash('Conta adicionada com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/pay_bill/<int:bill_id>', methods=['POST'])
@login_required
def handle_pay_bill(bill_id):
    # Agora recebemos o account_id do formulário do modal
    payment_account_id = request.form.get('payment_account_id', type=int) 
    
    if not payment_account_id:
        flash('Selecione uma conta para realizar o pagamento.', 'danger')
        return redirect(url_for('index'))

    if pay_bill_db(bill_id, current_user.id, payment_account_id):
        flash('Conta paga e transação registrada com sucesso!', 'success')
    else:
        # A função pay_bill_db já lida com saldo insuficiente e outras validações.
        # A flash message será gerada dentro dela.
        pass 
    return redirect(url_for('index'))

@app.route('/reschedule_bill/<int:bill_id>', methods=['POST'])
@login_required
def handle_reschedule_bill(bill_id):
    new_date = request.form['new_date']
    if reschedule_bill_db(bill_id, new_date, current_user.id):
        flash('Conta remarcada com sucesso!', 'success')
    else:
        flash('Não foi possível remarcar a conta. Verifique se ela existe ou pertence a você.', 'danger')
    return redirect(url_for('index'))

@app.route('/delete_transaction/<int:transaction_id>', methods=['POST'])
@login_required
def handle_delete_transaction(transaction_id):
    if delete_transaction_db(transaction_id, current_user.id):
        flash('Transação excluída com sucesso!', 'info')
    else:
        flash('Não foi possível excluir a transação. Verifique se ela existe ou pertence a você.', 'danger')
    return redirect(url_for('index'))

@app.route('/delete_bill/<int:bill_id>', methods=['POST'])
@login_required
def handle_delete_bill(bill_id):
    if delete_bill_db(bill_id, current_user.id):
        flash('Conta excluída com sucesso!', 'info')
    else:
        flash('Não foi possível excluir a conta. Verifique se ela existe ou pertence a você.', 'danger')
    return redirect(url_for('index'))

@app.route('/get_transaction_data/<int:transaction_id>', methods=['GET'])
@login_required
def get_transaction_data(transaction_id):
    transaction = Transaction.query.filter_by(id=transaction_id, user_id=current_user.id).first()
    if transaction:
        return jsonify({
            'id': transaction.id,
            'description': transaction.description,
            'amount': transaction.amount,
            'date': transaction.date,
            'type': transaction.type,
            'category_id': transaction.category_id,
            'account_id': transaction.account_id,
            'goal_id': transaction.goal_id
        })
    return jsonify({'error': 'Transação não encontrada ou não pertence a este usuário'}), 404

@app.route('/edit_transaction/<int:transaction_id>', methods=['POST'])
@login_required
def handle_edit_transaction(transaction_id):
    description = request.form['edit_description']
    amount = request.form['edit_amount']
    date = request.form['edit_date']
    transaction_type = request.form['edit_type']
    category_id = request.form.get('edit_category_id', type=int)
    account_id = request.form.get('edit_account_id', type=int)
    goal_id = request.form.get('edit_goal_id', type=int) if request.form.get('edit_goal_id') else None

    if edit_transaction_db(transaction_id, description, amount, date, transaction_type, current_user.id, category_id, account_id, goal_id):
        flash('Transação atualizada com sucesso!', 'success')
    else:
        flash('Não foi possível atualizar a transação. Verifique se ela existe ou pertence a você.', 'danger')
    return redirect(url_for('index'))

@app.route('/get_bill_data/<int:bill_id>', methods=['GET'])
@login_required
def get_bill_data(bill_id):
    bill = Bill.query.filter_by(id=bill_id, user_id=current_user.id).first()
    if bill:
        return jsonify({
            'id': bill.id,
            'description': bill.description,
            'amount': bill.amount,
            'dueDate': bill.dueDate,
            'status': bill.status,
            'is_master_recurring_bill': bill.is_master_recurring_bill,
            'recurring_parent_id': bill.recurring_parent_id,
            'recurring_child_number': bill.recurring_child_number,
            'recurring_frequency': bill.recurring_frequency,
            'recurring_start_date': bill.recurring_start_date,
            'recurring_next_due_date': bill.recurring_next_due_date,
            'recurring_total_occurrences': bill.recurring_total_occurrences,
            'recurring_installments_generated': bill.recurring_installments_generated,
            'is_active_recurring': bill.is_active_recurring,
            'type': bill.type,
            'category_id': bill.category_id,
            'account_id': bill.account_id
        })
    return jsonify({'error': 'Conta não encontrada ou não pertence a este usuário'}), 404

@app.route('/edit_bill/<int:bill_id>', methods=['POST'])
@login_required
def handle_edit_bill(bill_id):
    description = request.form['edit_bill_description']
    amount = request.form['edit_bill_amount']
    due_date = request.form['edit_bill_dueDate']
    
    is_recurring = request.form.get('edit_is_recurring_bill') == 'on'
    recurring_frequency = request.form.get('edit_recurring_frequency_bill')
    recurring_total_occurrences = request.form.get('edit_recurring_total_occurrences_bill', type=int)
    is_active_recurring = request.form.get('edit_is_active_recurring_bill') == 'on'
    bill_type = request.form['edit_bill_type']
    category_id = request.form.get('edit_bill_category_id', type=int)
    # account_id removido do formulário, então passamos o original ou None
    bill = Bill.query.filter_by(id=bill_id, user_id=current_user.id).first()
    account_id = bill.account_id if bill else None # Mantém o account_id original se não for editado via formulário

    if edit_bill_db(bill_id, description, amount, due_date, current_user.id,
                      is_recurring, recurring_frequency, recurring_total_occurrences, is_active_recurring, bill_type, category_id, account_id):
        flash('Conta atualizada com sucesso!', 'success')
    else:
        flash('Não foi possível atualizar a conta. Verifique se ela existe ou pertence a você.', 'danger')
    return redirect(url_for('index'))


# --- ROTAS DE AUTENTICAÇÃO E GERENCIAMENTO DE USUÁRIO ---

def create_default_data_for_user(user):
    # Cria categorias padrão
    db.session.add(Category(name='Salário', type='income', user_id=user.id))
    db.session.add(Category(name='Freelance', type='income', user_id=user.id))
    db.session.add(Category(name='Outras Receitas', type='income', user_id=user.id))
    db.session.add(Category(name='Alimentação', type='expense', user_id=user.id))
    db.session.add(Category(name='Transporte', type='expense', user_id=user.id))
    db.session.add(Category(name='Lazer', type='expense', user_id=user.id))
    db.session.add(Category(name='Moradia', type='expense', user_id=user.id))
    db.session.add(Category(name='Saúde', type='expense', user_id=user.id))
    db.session.add(Category(name='Educação', type='expense', user_id=user.id))
    db.session.add(Category(name='Contas Fixas', type='expense', user_id=user.id))
    db.session.add(Category(name='Outras Despesas', type='expense', user_id=user.id))
    db.session.add(Category(name='Poupança para Metas', type='expense', user_id=user.id))
    db.session.add(Category(name='Assinaturas', type='expense', user_id=user.id)) # NOVA CATEGORIA PADRÃO
    
    # Cria uma conta principal padrão
    db.session.add(Account(name='Conta Principal', balance=0.00, user_id=user.id))
    
    db.session.commit()
    print(f"Categorias e conta padrão criadas para o usuário '{user.email}'.")


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']

        if len(password) < 8:
            flash('A senha deve ter no mínimo 8 caracteres.', 'danger')
            return redirect(url_for('register'))

        existing_user_email = User.query.filter_by(email=email).first()
        existing_user_username = User.query.filter_by(username=username).first()

        if existing_user_email:
            flash('Este e-mail já está em uso. Por favor, escolha outro.', 'danger')
        elif existing_user_username:
            flash('Este nome de usuário já está em uso. Por favor, escolha outro.', 'danger')
        else:
            new_user = User(email=email, username=username)
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()

            create_default_data_for_user(new_user)

            flash('Conta criada com sucesso! Faça login.', 'success')
            return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        identifier = request.form['identifier']
        password = request.form['password']

        user = User.query.filter((User.email == identifier) | (User.username == identifier)).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Login realizado com sucesso!', 'success')
            return redirect(url_for('index'))
        else:
            flash('E-mail/Nome de usuário ou senha incorretos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você foi desconectado.', 'info')
    return redirect(url_for('login'))

@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html', current_user=current_user)

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form['old_password']
        new_password = request.form['new_password']
        confirm_new_password = request.form['confirm_new_password']

        if not current_user.check_password(old_password):
            flash('Senha antiga incorreta.', 'danger')
        elif len(new_password) < 8:
            flash('A nova senha deve ter no mínimo 8 caracteres.', 'danger')
        elif new_password != confirm_new_password:
            flash('A nova senha e a confirmação não coincidem.', 'danger')
        else:
            current_user.set_password(new_password)
            db.session.commit()
            flash('Senha alterada com sucesso!', 'success')
            return redirect(url_for('profile'))
    return render_template('change_password.html')

@app.route('/delete_account_user', methods=['GET', 'POST'])
@login_required
def delete_account_user():
    if request.method == 'POST':
        confirm_password = request.form['confirm_password']

        if current_user.check_password(confirm_password):
            user_to_delete = User.query.get(current_user.id)
            if user_to_delete:
                logout_user()
                db.session.delete(user_to_delete)
                db.session.commit()
                flash('Sua conta foi excluída permanentemente.', 'success')
                return redirect(url_for('register'))
            else:
                flash('Erro ao encontrar sua conta.', 'danger')
        else:
            flash('Senha incorreta.', 'danger')
    return render_template('delete_account.html')

# ROTA: Solicitar Código de Recuperação de Senha
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password_request():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        email = request.form['email']
        user = User.query.filter_by(email=email).first()

        if user:
            recovery_code = generate_recovery_code()
            user.recovery_code = recovery_code
            user.recovery_code_expires_at = datetime.datetime.now() + datetime.timedelta(minutes=10) # Código válido por 10 minutos
            db.session.commit()

            if send_recovery_email(user.email, recovery_code):
                flash('Um código de recuperação foi enviado para o seu e-mail.', 'success')
                return redirect(url_for('forgot_password_request', code_sent='true', email=email))
            else:
                flash('Não foi possível enviar o código de recuperação. Verifique suas configurações de e-mail.', 'danger')
        else:
            flash('E-mail não encontrado.', 'danger')
        
    return render_template('forgot_password.html')

# ROTA: Verificar Código e Redefinir Senha
@app.route('/reset_password_verify', methods=['POST'])
def reset_password_verify():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    email = request.form['email']
    code = request.form['code']
    new_password = request.form['new_password']
    confirm_new_password = request.form['confirm_new_password']

    user = User.query.filter_by(email=email).first()

    if not user:
        flash('E-mail não encontrado.', 'danger')
    elif user.recovery_code is None or user.recovery_code_expires_at is None:
        flash('Não há solicitação de recuperação de senha ativa para este e-mail. Solicite um novo código.', 'danger')
    elif user.recovery_code != code.upper(): # Converta para maiúsculas para comparar
        flash('Código de recuperação incorreto.', 'danger')
    elif datetime.datetime.now() > user.recovery_code_expires_at:
        flash('O código de recuperação expirou. Solicite um novo código.', 'danger')
    elif len(new_password) < 8:
        flash('A nova senha deve ter no mínimo 8 caracteres.', 'danger')
    elif new_password != confirm_new_password:
        flash('A nova senha e a confirmação não coincidem.', 'danger')
    else:
        user.set_password(new_password)
        user.recovery_code = None
        user.recovery_code_expires_at = None
        db.session.commit()
        flash('Sua senha foi redefinida com sucesso! Faça login.', 'success')
        return redirect(url_for('login'))
        
    # Se houver erro, renderiza a página de recuperação novamente com a seção de reset visível
    return render_template('forgot_password.html', code_sent='true', email=email)


# ROTA: Resumo Financeiro Mensal (para Profile)
@app.route('/profile/monthly_summary', methods=['GET'])
@login_required
def get_monthly_summary():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        current_date = datetime.date.today()
        year = current_date.year
        month = current_date.month

    start_date = datetime.date(year, month, 1)
    end_date = start_date.replace(day=calendar.monthrange(year, month)[1])
    
    start_date_str = start_date.isoformat()
    end_date_str = (end_date + datetime.timedelta(days=1)).isoformat()

    monthly_transactions = Transaction.query.filter(
        Transaction.user_id == current_user.id,
        Transaction.date >= start_date_str,
        Transaction.date < end_date_str
    ).all()

    monthly_income = sum(t.amount for t in monthly_transactions if t.type == 'income')
    monthly_expenses = sum(t.amount for t in monthly_transactions if t.type == 'expense')
    monthly_balance = monthly_income - monthly_expenses
    
    transactions_details = [
        {'description': t.description, 'amount': t.amount, 'type': t.type, 'date': t.date,
         'category': t.category.name if t.category else 'Sem Categoria'}
        for t in monthly_transactions
    ]
    
    return jsonify({
        'year': year,
        'month': month,
        'income': monthly_income,
        'expenses': monthly_expenses,
        'balance': monthly_balance,
        'transactions_details': transactions_details
    })

# ROTA: Insight da IA (para Profile e Relatórios)
@app.route('/ai_insight', methods=['POST'])
@login_required
def get_ai_insight():
    data = request.get_json()
    
    # Pode receber summary_data (do perfil) ou report_data (dos relatórios)
    summary_data = data.get('summary_data') 
    report_data = data.get('report_data')

    prompt_parts = []
    
    if summary_data:
        prompt_parts.extend([
            f"Com base nos seguintes dados financeiros de um mês específico para o usuário {current_user.username}:",
            f"Receita Total: R${summary_data['income']:.2f},",
            f"Despesa Total: R${summary_data['expenses']:.2f},",
            f"Saldo Mensal: R${summary_data['balance']:.2f}."
        ])
        if summary_data.get('transactions_details'):
            limited_transactions = summary_data['transactions_details'][:5]
            trans_str = ", ".join([f"{t['description']} (R${t['amount']:.2f}, {t['type']}, {t['category']})" for t in limited_transactions])
            prompt_parts.append(f"Principais transações: {trans_str}.")

    elif report_data:
        start_date = report_data.get('start_date', 'N/A')
        end_date = report_data.get('end_date', 'N/A')
        total_income = report_data.get('total_income', 0.0)
        total_expenses = report_data.get('total_expenses', 0.0)
        balance = report_data.get('balance', 0.0)
        expenses_by_category = report_data.get('expenses_by_category', {})
        transaction_count = report_data.get('transaction_count', 0)

        prompt_parts.extend([
            f"Com base nos dados financeiros do usuário {current_user.username} para o período de {start_date} a {end_date}:",
            f"Receita Total: R${total_income:.2f},",
            f"Despesa Total: R${total_expenses:.2f},",
            f"Saldo Líquido no Período: R${balance:.2f}.",
            f"Total de {transaction_count} transações registradas."
        ])
        if expenses_by_category:
            expense_cat_str = ", ".join([f"{cat}: R${val:.2f}" for cat, val in expenses_by_category.items()])
            prompt_parts.append(f"Despesas por Categoria: {expense_cat_str}.")
        
    else:
        return jsonify({'error': 'Dados de resumo ou relatório não fornecidos.'}), 400

    prompt_parts.append(
        f"Forneça um breve insight ou conselho financeiro pessoal para o usuário. "
        f"Concentre-se em pontos fortes, áreas para melhoria ou tendências. "
        f"Use uma linguagem amigável e direta em português. "
        f"Seja conciso, com no máximo 150 palavras (pode exceder ligeiramente se necessário para clareza). "
        f"Não inclua 'Olá!' ou saudações, vá direto ao ponto."
    )

    prompt_text = " ".join(prompt_parts)
    print(f"DEBUG: Prompt enviado ao Gemini: {prompt_text}")
    ai_text = generate_text_with_gemini(prompt_text)
    return jsonify({'insight': ai_text})


@app.route('/profile/update_picture', methods=['POST'])
@login_required
def update_profile_picture():
    picture_url = request.form['profile_picture_url']
    current_user.profile_picture_url = picture_url
    db.session.commit()
    flash('Foto de perfil atualizada com sucesso!', 'success')
    return redirect(url_for('profile'))

@app.route('/get_chart_data', methods=['GET'])
@login_required
def get_chart_data():
    user_id = current_user.id
    
    today = datetime.date.today()
    
    month_labels = []
    monthly_income_data = {}
    monthly_expenses_data = {}

    for i in range(6, -1, -1): # Últimos 7 meses (mês atual + 6 anteriores)
        target_month_date = today - relativedelta(months=i)
        target_month = target_month_date.month
        target_year = target_month_date.year
        
        month_name = datetime.date(target_year, target_month, 1).strftime('%b/%Y')
        month_labels.append(month_name)
        
        start_date_of_month = target_month_date.replace(day=1)
        end_date_of_month = target_month_date.replace(day=calendar.monthrange(target_year, target_month)[1])

        start_date_str = start_date_of_month.isoformat()
        end_date_str = (end_date_of_month + datetime.timedelta(days=1)).isoformat()

        transactions_in_month = Transaction.query.filter(
            Transaction.user_id == user_id,
            Transaction.date >= start_date_str,
            Transaction.date < end_date_str
        ).all()
        
        monthly_income_data[month_name] = sum(t.amount for t in transactions_in_month if t.type == 'income')
        monthly_expenses_data[month_name] = sum(t.amount for t in transactions_in_month if t.type == 'expense')

    monthly_overview_chart_data = {
        'labels': month_labels,
        'income': [monthly_income_data[m] for m in month_labels],
        'expenses': [monthly_expenses_data[m] for m in month_labels]
    }

    expenses_by_category = {}
    
    current_year_start = today.replace(month=1, day=1)
    current_year_end = today.replace(month=12, day=31)

    start_year_str = current_year_start.isoformat()
    end_year_str = (current_year_end + datetime.timedelta(days=1)).isoformat()

    current_year_transactions = Transaction.query.filter(
        Transaction.user_id == user_id,
        Transaction.type == 'expense', # Apenas despesas para este gráfico
        Transaction.date >= start_year_str,
        Transaction.date < end_year_str
    ).all()

    for transaction in current_year_transactions:
        category_name = transaction.category.name if transaction.category else 'Sem Categoria'
        expenses_by_category[category_name] = expenses_by_category.get(category_name, 0) + transaction.amount

    expenses_by_category_chart_data = {
        'labels': list(expenses_by_category.keys()),
        'values': list(expenses_by_category.values())
    }

    return jsonify({
        'monthly_summary': monthly_overview_chart_data,
        'expenses_by_category': expenses_by_category_chart_data
    })

@app.route('/budgets')
@login_required
def budgets_page():
    user_id = current_user.id
    
    selected_month_year = request.args.get('month_year', get_current_month_year_str())
    
    budgets = Budget.query.filter_by(user_id=user_id, month_year=selected_month_year).all()
    expense_categories = Category.query.filter_by(user_id=user_id, type='expense').all()

    start_date_obj, end_date_obj = get_month_start_end_dates(selected_month_year)
    start_date_str = start_date_obj.isoformat()
    end_date_str = end_date_obj.isoformat()

    for budget in budgets:
        total_spent_in_category = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.user_id == user_id,
            Transaction.category_id == budget.category_id,
            Transaction.type == 'expense',
            Transaction.date >= start_date_str,
            Transaction.date <= end_date_str
        ).scalar() or 0.0
        budget.current_spent = total_spent_in_category

    current_date = datetime.datetime.strptime(selected_month_year + '-01', '%Y-%m-%d').date()
    prev_month = (current_date - relativedelta(months=1)).strftime('%Y-%m')
    next_month = (current_date + relativedelta(months=1)).strftime('%Y-%m')

    return render_template('budgets.html', 
                           budgets=budgets, 
                           expense_categories=expense_categories, 
                           current_month_year=selected_month_year,
                           prev_month_year=prev_month,
                           next_month_year=next_month)

@app.route('/add_budget', methods=['POST'])
@login_required
def handle_add_budget():
    category_id = request.form['category_id']
    budget_amount = request.form['budget_amount']
    month_year = request.form['month_year']

    if add_budget_db(current_user.id, category_id, budget_amount, month_year):
        flash('Orçamento adicionado/atualizado com sucesso!', 'success')
    else:
        flash('Erro ao adicionar/atualizar orçamento. Um orçamento para esta categoria e mês já pode existir.', 'danger')
    return redirect(url_for('budgets_page', month_year=month_year))

@app.route('/edit_budget/<int:budget_id>', methods=['POST'])
@login_required
def handle_edit_budget(budget_id):
    budget_amount = request.form['budget_amount']
    if edit_budget_db(budget_id, current_user.id, budget_amount):
        flash('Orçamento atualizado com sucesso!', 'success')
    else:
        flash('Erro ao atualizar orçamento.', 'danger')
    budget = Budget.query.get(budget_id)
    redirect_month_year = budget.month_year if budget else get_current_month_year_str()
    return redirect(url_for('budgets_page', month_year=redirect_month_year))

@app.route('/delete_budget/<int:budget_id>', methods=['POST'])
@login_required
def handle_delete_budget(budget_id):
    if delete_budget_db(budget_id, current_user.id):
        flash('Orçamento excluído com sucesso!', 'info')
    else:
        flash('Erro ao excluir orçamento.', 'danger')
    redirect_month_year = get_current_month_year_str() 
    return redirect(url_for('budgets_page', month_year=redirect_month_year))

@app.route('/goals', methods=['GET'])
@login_required
def goals_page():
    goals = Goal.query.filter_by(user_id=current_user.id).all()
    user_accounts = Account.query.filter_by(user_id=current_user.id).all() # Passa as contas para o template
    accounts_json = [{'id': acc.id, 'name': acc.name, 'balance': acc.balance} for acc in user_accounts]
    return render_template('goals.html', goals=goals, accounts=user_accounts, accounts_json=accounts_json) # Passa accounts_json

@app.route('/add_goal', methods=['POST'])
@login_required
def handle_add_goal():
    name = request.form['name']
    target_amount = request.form['target_amount']
    due_date = request.form['due_date'] if request.form['due_date'] else None

    if add_goal_db(current_user.id, name, target_amount, due_date):
        flash('Meta adicionada com sucesso!', 'success')
    else:
        flash('Erro ao adicionar meta.', 'danger')
    return redirect(url_for('goals_page'))

@app.route('/edit_goal/<int:goal_id>', methods=['POST'])
@login_required
def handle_edit_goal(goal_id):
    name = request.form['name']
    target_amount = request.form['target_amount']
    current_amount = request.form['current_amount']
    due_date = request.form['due_date'] if request.form['due_date'] else None
    status = request.form['status']

    if edit_goal_db(goal_id, current_user.id, name, target_amount, current_amount, due_date, status):
        flash('Meta atualizada com sucesso!', 'success')
    else:
        flash('Erro ao atualizar meta.', 'danger')
    return redirect(url_for('goals_page'))

@app.route('/delete_goal/<int:goal_id>', methods=['POST'])
@login_required
def handle_delete_goal(goal_id):
    if delete_goal_db(goal_id, current_user.id):
        flash('Meta excluída com sucesso!', 'info')
    else:
        flash('Erro ao excluir meta.', 'danger')
    return redirect(url_for('goals_page'))

@app.route('/contribute_to_goal/<int:goal_id>', methods=['POST'])
@login_required
def handle_contribute_to_goal(goal_id):
    amount = request.form['amount']
    source_account_id = request.form.get('source_account_id', type=int)

    if not source_account_id:
        flash('Selecione uma conta de origem para a contribuição.', 'danger')
        return redirect(url_for('goals_page'))

    if contribute_to_goal_db(goal_id, current_user.id, amount, source_account_id):
        pass 
    else:
        flash('Erro ao contribuir para a meta. Verifique o valor, a conta ou se a meta existe.', 'danger')
    return redirect(url_for('goals_page'))


# --- NOVAS ROTAS PARA ORÇAMENTO INTELIGENTE ---
@app.route('/recreate_budget')
@login_required
def recreate_last_month_budget():
    current_month_date = datetime.date.today().replace(day=1)
    last_month_date = current_month_date - relativedelta(months=1)
    last_month_year = last_month_date.strftime('%Y-%m')
    current_month_year = current_month_date.strftime('%Y-%m')

    last_month_budgets = Budget.query.filter_by(user_id=current_user.id, month_year=last_month_year).all()

    if not last_month_budgets:
        flash('Nenhum orçamento encontrado no mês anterior para copiar.', 'warning')
        return redirect(url_for('budgets_page'))

    for budget in last_month_budgets:
        add_budget_db(current_user.id, budget.category_id, budget.budget_amount, current_month_year)
    
    flash('Orçamento do mês anterior recriado com sucesso!', 'success')
    return redirect(url_for('budgets_page'))

@app.route('/suggest_budget_ai')
@login_required
def suggest_budget_ai(): # Renomeado para evitar conflito com a função interna
    current_month_date = datetime.date.today().replace(day=1)
    last_month_date = current_month_date - relativedelta(months=1)
    last_month_year = last_month_date.strftime('%Y-%m')
    current_month_year = current_month_date.strftime('%Y-%m')

    last_month_budgets = Budget.query.filter_by(user_id=current_user.id, month_year=last_month_year).all()

    if not last_month_budgets:
        flash('Nenhum orçamento encontrado no mês anterior para usar como base para a sugestão.', 'warning')
        return redirect(url_for('budgets_page'))

    # Coleta dados de gastos reais do mês anterior
    budget_vs_actual = []
    for budget in last_month_budgets:
        budget_vs_actual.append(
            f"- Categoria: {budget.category.name}, Orçado: R${budget.budget_amount:.2f}, Gasto Real: R${budget.current_spent:.2f}"
        )
    
    data_summary = "\n".join(budget_vs_actual)

    prompt = (
        f"Você é um assistente financeiro. Com base nos seguintes dados financeiros de um mês específico para o usuário {current_user.username}:" # Usa username para IA
        f"sugira um novo orçamento para o mês atual. Ajuste os valores de forma realista. "
        f"Se o gasto foi muito maior que o orçado, sugira um aumento moderado. Se foi muito menor, sugira uma pequena redução. "
        f"Mantenha os valores arredondados para facilitar. Responda APENAS com um JSON contendo uma chave 'sugestoes' que é uma lista de objetos, "
        f"onde cada objeto tem as chaves 'categoria' e 'valor_sugerido'.\n\n"
        f"Dados do Mês Anterior:\n{data_summary}"
    )

    try:
        ai_response_text = generate_text_with_gemini(prompt).replace("```json", "").replace("```", "").strip()
        suggestions = json.loads(ai_response_text)

        if 'sugestoes' not in suggestions:
            raise ValueError("Resposta da IA não contém a chave 'sugestoes'.")

        for sug in suggestions['sugestoes']:
            category = Category.query.filter_by(user_id=current_user.id, name=sug['categoria'], type='expense').first()
            if category:
                add_budget_db(current_user.id, category.id, sug['valor_sugerido'], current_month_year)
        
        flash('Orçamento sugerido pela IA foi criado! Revise e ajuste se necessário.', 'success')
    except Exception as e:
        print(f"ERROR: Erro ao processar sugestão da IA: {e}")
        flash('Não foi possível gerar uma sugestão da IA no momento. Por favor, tente novamente mais tarde.', 'danger')

    return redirect(url_for('budgets_page'))

# --- ROTAS DE GERENCIAMENTO DE CONTAS ---
@app.route('/accounts')
@login_required
def accounts_page():
    """Exibe a página de gerenciamento de contas."""
    user_accounts = Account.query.filter_by(user_id=current_user.id).order_by(Account.name.asc()).all()
    accounts_json = [{'id': acc.id, 'name': acc.name, 'balance': acc.balance} for acc in user_accounts]
    return render_template('accounts.html', accounts=user_accounts, accounts_json=accounts_json)

@app.route('/add_account', methods=['POST'])
@login_required
def handle_add_account():
    """Lida com a adição de uma nova conta."""
    name = request.form['name']
    initial_balance = request.form['initial_balance']
    
    if add_account_db(current_user.id, name, initial_balance):
        flash('Conta adicionada com sucesso!', 'success')
    else:
        flash('Erro ao adicionar conta. Uma conta com este nome já pode existir.', 'danger')
    return redirect(url_for('accounts_page'))

@app.route('/edit_account/<int:account_id>', methods=['POST'])
@login_required
def handle_edit_account(account_id):
    """Lida com a edição de uma conta existente."""
    name = request.form['name']
    balance = request.form['balance']
    
    if edit_account_db(account_id, current_user.id, name, balance):
        flash('Conta atualizada com sucesso!', 'success')
    else:
        flash('Erro ao atualizar conta. Uma conta com este nome já pode existir.', 'danger')
    return redirect(url_for('accounts_page'))

@app.route('/delete_account/<int:account_id>', methods=['POST'])
@login_required
def handle_delete_account(account_id):
    """Lida com a exclusão de uma conta."""
    if delete_account_db(account_id, current_user.id):
        flash('Conta excluída com sucesso! Transações associadas foram desvinculadas.', 'info')
    else:
        flash('Erro ao excluir conta. Verifique se ela existe ou pertence a você.', 'danger')
    return redirect(url_for('accounts_page'))

@app.route('/get_account_data/<int:account_id>', methods=['GET'])
@login_required
def get_account_data(account_id):
    """Retorna dados de uma conta específica para edição via AJAX."""
    account = Account.query.filter_by(id=account_id, user_id=current_user.id).first()
    if account:
        return jsonify({
            'id': account.id,
            'name': account.name,
            'balance': account.balance
        })
    return jsonify({'error': 'Conta não encontrada ou não pertence a este usuário'}), 404

@app.route('/transfer_funds', methods=['POST'])
@login_required
def handle_transfer_funds():
    """Lida com a transferência de fundos entre contas."""
    source_account_id = request.form.get('source_account_id', type=int)
    destination_account_id = request.form.get('destination_account_id', type=int)
    amount = request.form.get('amount', type=float)

    if transfer_funds_db(current_user.id, source_account_id, destination_account_id, amount):
        pass
    else:
        flash('Erro ao realizar a transferência.', 'danger')
    return redirect(url_for('accounts_page'))

# --- ROTAS DE RELATÓRIOS ---
@app.route('/reports')
@login_required
def reports_page():
    """Exibe a página de relatórios."""
    all_categories_formatted = [(c.id, c.type, c.name) for c in Category.query.filter_by(user_id=current_user.id).all()]
    return render_template('reports.html', all_categories=all_categories_formatted)

@app.route('/get_detailed_report_data', methods=['GET'])
@login_required
def get_detailed_report_data():
    """
    Endpoint para buscar dados detalhados para relatórios com base em filtros.
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    transaction_type = request.args.get('transaction_type')
    category_id = request.args.get('category_id', type=int)

    if not start_date_str or not end_date_str:
        return jsonify({'error': 'Datas de início e fim são obrigató.'}), 400

    report_data = get_detailed_report_data_db(current_user.id, start_date_str, end_date_str, transaction_type, category_id)
    
    if 'error' in report_data:
        return jsonify(report_data), 400
    
    return jsonify(report_data)

@app.route('/export_report/<format>', methods=['GET'])
@login_required
def export_report(format):
    """
    Exporta os dados do relatório para Excel ou PDF com base nos filtros.
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    transaction_type = request.args.get('transaction_type')
    category_id = request.args.get('category_id', type=int)

    if not start_date_str or not end_date_str:
        flash('Datas de início e fim são obrigatórias para exportação.', 'danger')
        return redirect(url_for('reports_page'))

    # Reutilize a função que busca os dados do relatório
    report_data = get_detailed_report_data_db(current_user.id, start_date_str, end_date_str, transaction_type, category_id)

    if 'error' in report_data:
        flash(f"Erro ao gerar dados para exportação: {report_data['error']}", 'danger')
        return redirect(url_for('reports_page'))

    transactions_query = Transaction.query.filter(
        Transaction.user_id == current_user.id,
        Transaction.date >= start_date_str,
        Transaction.date <= end_date_str
    )
    if transaction_type and transaction_type in ['income', 'expense']:
        transactions_query = transactions_query.filter(Transaction.type == transaction_type)
    if category_id:
        transactions_query = transactions_query.filter(Transaction.category_id == category_id)
    
    filtered_transactions = transactions_query.order_by(Transaction.date.asc()).all()

    if format == 'excel':
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatorio Financeiro"

        # Cabeçalho
        headers = ["Descrição", "Valor", "Data", "Tipo", "Categoria", "Conta", "Meta"]
        sheet.append(headers)

        # Estilo para o cabeçalho
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light gray
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        # Dados das transações
        for row_idx, t in enumerate(filtered_transactions, start=2): # Começa da linha 2
            category_name = t.category.name if t.category else "N/A"
            account_name = t.account.name if t.account else "N/A"
            goal_name = t.goal.name if t.goal else "N/A"
            row_data = [
                t.description,
                t.amount,
                t.date,
                "Receita" if t.type == "income" else "Despesa",
                category_name,
                account_name,
                goal_name
            ]
            sheet.append(row_data)

            # Aplicar formato de moeda à coluna 'Valor' (coluna B)
            sheet.cell(row=row_idx, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE # Formato de moeda

        # Auto-ajustar largura das colunas
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter # Get the column name (e.g. 'A')
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar resumo de despesas por categoria
        if report_data['expenses_by_category_chart']['labels']:
            sheet.append([]) # Linha em branco para separação
            sheet.append(["Despesas por Categoria"])
            for i, label in enumerate(report_data['expenses_by_category_chart']['labels']):
                value = report_data['expenses_by_category_chart']['values'][i]
                sheet.append([label, value])
                # A célula de valor está na coluna 2 da linha recém-adicionada
                sheet.cell(row=sheet.max_row, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE # Formato de moeda

        # Adicionar evolução do patrimônio líquido
        if report_data['net_worth_evolution_chart']['labels']:
            sheet.append([]) # Linha em branco para separação
            sheet.append(["Evolução do Patrimônio Líquido"])
            sheet.append(["Data", "Patrimônio Líquido"])
            for i, label in enumerate(report_data['net_worth_evolution_chart']['labels']):
                value = report_data['net_worth_evolution_chart']['values'][i]
                # Filtrar valores: apenas se o valor for > 0.01 ou < 0.00
                if value > 0.01 or value < 0.00:
                    sheet.append([label, value])
                    # A célula de valor está na coluna 2 da linha recém-adicionada
                    sheet.cell(row=sheet.max_row, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE # Formato de moeda

        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name="relatorio_financeiro.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif format == 'pdf':
        # Configuração do PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size = 12)

        pdf.cell(200, 10, text = f"Relatório Financeiro de {start_date_str} a {end_date_str}", new_x="LMARGIN", new_y="NEXT", align = 'C')
        pdf.ln(10)

        # Detalhes das Transações
        if filtered_transactions:
            pdf.set_font("Arial", size = 10, style='B')
            pdf.cell(0, 10, text="Transações Detalhadas:", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Arial", size = 8)
            
            # Cabeçalho da tabela
            # Ajuste as larguras conforme necessário. Aumentado a largura da descrição.
            col_widths = [60, 20, 20, 20, 40, 30] 
            pdf.cell(col_widths[0], 7, "Descrição", 1)
            pdf.cell(col_widths[1], 7, "Valor", 1)
            pdf.cell(col_widths[2], 7, "Data", 1)
            pdf.cell(col_widths[3], 7, "Tipo", 1)
            pdf.cell(col_widths[4], 7, "Categoria", 1)
            pdf.cell(col_widths[5], 7, "Conta", 1)
            pdf.ln()

            for t in filtered_transactions:
                category_name = t.category.name if t.category else "N/A"
                account_name = t.account.name if t.account else "N/A"
                # Usar multi_cell para descrições longas
                # pdf.multi_cell(col_widths[0], 7, t.description, 1) # Isso faria a linha quebrar
                
                # Para evitar quebra de linha e manter a célula na mesma linha,
                # vamos truncar a descrição se for muito longa e garantir que não haja quebras de linha
                description_display = t.description.replace('\n', ' ').replace('\r', ' ')
                if len(description_display) > 35: # Limite aproximado para a largura da célula
                    description_display = description_display[:32] + "..."

                pdf.cell(col_widths[0], 7, description_display, 1)
                pdf.cell(col_widths[1], 7, f"R$ {t.amount:.2f}", 1)
                pdf.cell(col_widths[2], 7, t.date, 1)
                pdf.cell(col_widths[3], 7, "Receita" if t.type == "income" else "Despesa", 1)
                pdf.cell(col_widths[4], 7, category_name, 1)
                pdf.cell(col_widths[5], 7, account_name, 1)
                pdf.ln()
        else:
            pdf.cell(0, 10, text="Nenhuma transação encontrada para os filtros selecionados.", new_x="LMARGIN", new_y="NEXT")
        
        pdf.ln(10)

        # Resumo de Despesas por Categoria
        if report_data['expenses_by_category_chart']['labels']:
            pdf.set_font("Arial", size = 10, style='B')
            pdf.cell(0, 10, text="Despesas por Categoria:", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Arial", size = 8)
            for i, label in enumerate(report_data['expenses_by_category_chart']['labels']):
                value = report_data['expenses_by_category_chart']['values'][i]
                pdf.cell(0, 7, text=f"{label}: R$ {value:.2f}", new_x="LMARGIN", new_y="NEXT")
        
        pdf.ln(10)

        # Evolução do Patrimônio Líquido
        if report_data['net_worth_evolution_chart']['labels']:
            pdf.set_font("Arial", size = 10, style='B')
            pdf.cell(0, 10, text="Evolução do Patrimônio Líquido:", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Arial", size = 8)
            for i, label in enumerate(report_data['net_worth_evolution_chart']['labels']):
                value = report_data['net_worth_evolution_chart']['values'][i]
                # Filtrar valores: apenas se o valor for > 0.01 ou < 0.00
                if value > 0.01 or value < 0.00:
                    pdf.cell(0, 7, text=f"{label}: R$ {value:.2f}", new_x="LMARGIN", new_y="NEXT")

        # Correção aqui: remove .encode('latin-1')
        return send_file(io.BytesIO(pdf.output(dest='S')), download_name="relatorio_financeiro.pdf", as_attachment=True, mimetype='application/pdf')

    else:
        flash('Formato de exportação inválido.', 'danger')
        return redirect(url_for('reports_page'))

# --- ROTAS PARA GERENCIAMENTO DE ASSINATURAS ---
@app.route('/subscriptions')
@login_required
def subscriptions_page():
    user_id = current_user.id
    subscriptions = Subscription.query.filter_by(user_id=user_id).all()
    expense_categories = Category.query.filter_by(user_id=user_id, type='expense').all()
    user_accounts = Account.query.filter_by(user_id=user_id).all()
    return render_template('subscriptions.html', 
                           subscriptions=subscriptions, 
                           expense_categories=expense_categories,
                           accounts=user_accounts)

@app.route('/add_subscription', methods=['POST'])
@login_required
def handle_add_subscription():
    name = request.form['name']
    amount = request.form['amount']
    billing_cycle = request.form['billing_cycle']
    due_date_of_month = request.form['due_date_of_month']
    category_id = request.form.get('category_id', type=int)
    account_id = request.form.get('account_id', type=int)

    # Calcular a próxima data de vencimento inicial
    today = datetime.date.today()
    next_due_date_obj = None
    try:
        due_day = int(due_date_of_month)
        if not (1 <= due_day <= 31):
            flash('Dia de vencimento inválido. Deve ser entre 1 e 31.', 'danger')
            return redirect(url_for('subscriptions_page'))
        
        # Tenta criar a data no mês atual
        try:
            next_due_date_obj = today.replace(day=due_day)
        except ValueError: # Dia inválido para o mês atual (ex: 31 de fev)
            # Tenta o último dia do mês
            last_day_of_month = calendar.monthrange(today.year, today.month)[1]
            next_due_date_obj = today.replace(day=last_day_of_month)

        # Se a data de vencimento já passou no mês atual, define para o próximo mês
        if next_due_date_obj < today:
            next_due_date_obj += relativedelta(months=1)
            # Garante que o dia ainda é válido para o novo mês
            try:
                next_due_date_obj = next_due_date_obj.replace(day=due_day)
            except ValueError:
                last_day_of_next_month = calendar.monthrange(next_due_date_obj.year, next_due_date_obj.month)[1]
                next_due_date_obj = next_due_date_obj.replace(day=last_day_of_next_month)

    except ValueError:
        flash('Dia de vencimento inválido. Insira um número.', 'danger')
        return redirect(url_for('subscriptions_page'))
    
    if not next_due_date_obj:
        flash('Não foi possível determinar a próxima data de vencimento.', 'danger')
        return redirect(url_for('subscriptions_page'))

    new_subscription = Subscription(
        user_id=current_user.id,
        name=name,
        amount=float(amount),
        billing_cycle=billing_cycle,
        due_date_of_month=int(due_date_of_month),
        next_due_date=next_due_date_obj.isoformat(),
        status='active',
        category_id=category_id,
        account_id=account_id
    )
    db.session.add(new_subscription)
    db.session.commit()
    flash('Assinatura adicionada com sucesso!', 'success')
    return redirect(url_for('subscriptions_page'))

@app.route('/get_subscription_data/<int:subscription_id>', methods=['GET'])
@login_required
def get_subscription_data(subscription_id):
    subscription = Subscription.query.filter_by(id=subscription_id, user_id=current_user.id).first()
    if subscription:
        return jsonify({
            'id': subscription.id,
            'name': subscription.name,
            'amount': subscription.amount,
            'billing_cycle': subscription.billing_cycle,
            'due_date_of_month': subscription.due_date_of_month,
            'next_due_date': subscription.next_due_date,
            'status': subscription.status,
            'category_id': subscription.category_id,
            'account_id': subscription.account_id
        })
    return jsonify({'error': 'Assinatura não encontrada ou não pertence a este usuário'}), 404

@app.route('/edit_subscription/<int:subscription_id>', methods=['POST'])
@login_required
def handle_edit_subscription(subscription_id):
    subscription = Subscription.query.filter_by(id=subscription_id, user_id=current_user.id).first()
    if not subscription:
        flash('Assinatura não encontrada.', 'danger')
        return redirect(url_for('subscriptions_page'))

    name = request.form['edit_name']
    amount = request.form['edit_amount']
    billing_cycle = request.form['edit_billing_cycle']
    due_date_of_month = request.form['edit_due_date_of_month']
    status = request.form['edit_status']
    category_id = request.form.get('edit_category_id', type=int)
    account_id = request.form.get('edit_account_id', type=int)

    # Atualizar campos
    subscription.name = name
    subscription.amount = float(amount)
    subscription.billing_cycle = billing_cycle
    subscription.status = status
    subscription.category_id = category_id
    subscription.account_id = account_id

    # Recalcular next_due_date se o dia do mês ou ciclo de cobrança mudar
    if int(due_date_of_month) != subscription.due_date_of_month or billing_cycle != subscription.billing_cycle:
        subscription.due_date_of_month = int(due_date_of_month)
        
        # Lógica para recalcular next_due_date (similar à adição)
        today = datetime.date.today()
        next_due_date_obj = None
        try:
            due_day = int(due_date_of_month)
            if not (1 <= due_day <= 31):
                flash('Dia de vencimento inválido. Deve ser entre 1 e 31.', 'danger')
                return redirect(url_for('subscriptions_page'))
            
            # Tenta criar a data no mês atual
            try:
                next_due_date_obj = today.replace(day=due_day)
            except ValueError: # Dia inválido para o mês atual (ex: 31 de fev)
                last_day_of_month = calendar.monthrange(today.year, today.month)[1]
                next_due_date_obj = today.replace(day=last_day_of_month)

            # Se a data de vencimento já passou no mês atual, define para o próximo mês
            if next_due_date_obj < today:
                next_due_date_obj += relativedelta(months=1)
                # Garante que o dia ainda é válido para o novo mês
                try:
                    next_due_date_obj = next_due_date_obj.replace(day=due_day)
                except ValueError:
                    last_day_of_next_month = calendar.monthrange(next_due_date_obj.year, next_due_date_obj.month)[1]
                    next_due_date_obj = next_due_date_obj.replace(day=last_day_of_next_month)

        except ValueError:
            flash('Dia de vencimento inválido. Insira um número.', 'danger')
            return redirect(url_for('subscriptions_page'))
        
        if next_due_date_obj:
            subscription.next_due_date = next_due_date_obj.isoformat()
        else:
            flash('Não foi possível recalcular a próxima data de vencimento.', 'danger')
            return redirect(url_for('subscriptions_page'))

    db.session.commit()
    flash('Assinatura atualizada com sucesso!', 'success')
    return redirect(url_for('subscriptions_page'))

@app.route('/delete_subscription/<int:subscription_id>', methods=['POST'])
@login_required
def handle_delete_subscription(subscription_id):
    subscription = Subscription.query.filter_by(id=subscription_id, user_id=current_user.id).first()
    if not subscription:
        flash('Assinatura não encontrada.', 'danger')
        return redirect(url_for('subscriptions_page'))
    
    db.session.delete(subscription)
    db.session.commit()
    flash('Assinatura excluída com sucesso!', 'info')
    return redirect(url_for('subscriptions_page'))


# --- INICIALIZAÇÃO DO BANCO DE DADOS ---
with app.app_context():
    db.init_app(app) # Associar SQLAlchemy ao app
    migrate.init_app(app, db) # Associar Flask-Migrate ao app e db
    db.create_all() # Cria as tabelas se não existirem (para SQLite local ou primeira vez)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
