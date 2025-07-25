from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user

# Importa as funções de serviço e os modelos
from services import get_detailed_report_data_db, generate_text_with_gemini
from models import Category, Transaction # Importa Transaction para a lista filtrada

report_bp = Blueprint('report', __name__)

@report_bp.route('/reports')
@login_required
def reports_page():
    """Exibe a página de relatórios."""
    # Obtém todas as categorias do usuário para o filtro
    all_categories_formatted = [(c.id, c.type, c.name) for c in Category.query.filter_by(user_id=current_user.id).all()]
    return render_template('reports.html', all_categories=all_categories_formatted)

@report_bp.route('/get_detailed_report_data', methods=['GET'])
@login_required
def get_detailed_report_data():
    """
    Endpoint para buscar dados detalhados para relatórios com base em filtros via AJAX.
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    transaction_type = request.args.get('transaction_type')
    category_id = request.args.get('category_id', type=int)

    if not start_date_str or not end_date_str:
        return jsonify({'error': 'Datas de início e fim são obrigatórias.'}), 400

    report_data = get_detailed_report_data_db(current_user.id, start_date_str, end_date_str, transaction_type, category_id)
    
    if 'error' in report_data:
        return jsonify(report_data), 400
    
    return jsonify(report_data)

@report_bp.route('/export_report/<format>', methods=['GET'])
@login_required
def export_report(format):
    """
    Exporta os dados do relatório para Excel ou PDF com base nos filtros.
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    transaction_type = request.args.get('transaction_type')
    category_id = request.args.get('category_id', type=int)

    if not start_date_str or not end_date_str:
        flash('Datas de início e fim são obrigatórias para exportação.', 'danger')
        return redirect(url_for('report.reports_page'))

    # Reutiliza a função que busca os dados do relatório (inclui o resumo para IA)
    report_data = get_detailed_report_data_db(current_user.id, start_date_str, end_date_str, transaction_type, category_id)

    if 'error' in report_data:
        flash(f"Erro ao gerar dados para exportação: {report_data['error']}", 'danger')
        return redirect(url_for('report.reports_page'))

    # Busca as transações filtradas para incluir no relatório exportado
    transactions_query = Transaction.query.filter(
        Transaction.user_id == current_user.id,
        Transaction.date >= start_date_str,
        Transaction.date <= end_date_str
    )
    if transaction_type and transaction_type in ['income', 'expense']:
        transactions_query = transactions_query.filter(Transaction.type == transaction_type)
    if category_id:
        transactions_query = transactions_query.filter(Transaction.category_id == category_id)
    
    filtered_transactions = transactions_query.order_by(Transaction.date.asc()).all()

    if format == 'excel':
        # Importa apenas quando necessário para evitar dependências cíclicas
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
        import io

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatorio Financeiro"

        # Cabeçalho
        headers = ["Descrição", "Valor", "Data", "Tipo", "Categoria", "Conta", "Meta"]
        sheet.append(headers)

        # Estilo para o cabeçalho
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light gray
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        
        # Dados das transações
        for row_idx, t in enumerate(filtered_transactions, start=2): # Começa da linha 2
            category_name = t.category.name if t.category else "N/A"
            account_name = t.account.name if t.account else "N/A"
            goal_name = t.goal.name if t.goal else "N/A"
            row_data = [
                t.description,
                t.amount,
                t.date,
                "Receita" if t.type == "income" else "Despesa",
                category_name,
                account_name,
                goal_name
            ]
            sheet.append(row_data)

            # Aplicar formato de moeda à coluna 'Valor' (coluna B)
            sheet.cell(row=row_idx, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE # Formato de moeda

        # Auto-ajustar largura das colunas
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter # Obtém a letra da coluna (ex: 'A')
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar resumo de despesas por categoria
        if report_data['expenses_by_category_chart']['labels']:
            sheet.append([]) # Linha em branco para separação
            sheet.append(["Despesas por Categoria"])
            for i, label in enumerate(report_data['expenses_by_category_chart']['labels']):
                value = report_data['expenses_by_category_chart']['values'][i]
                sheet.append([label, value])
                # A célula de valor está na coluna 2 da linha recém-adicionada
                sheet.cell(row=sheet.max_row, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE # Formato de moeda

        # Adicionar evolução do patrimônio líquido
        if report_data['net_worth_evolution_chart']['labels']:
            sheet.append([]) # Linha em branco para separação
            sheet.append(["Evolução do Patrimônio Líquido"])
            sheet.append(["Data", "Patrimônio Líquido"])
            for i, label in enumerate(report_data['net_worth_evolution_chart']['labels']):
                value = report_data['net_worth_evolution_chart']['values'][i]
                # Filtrar valores: apenas se o valor for > 0.01 ou < 0.00 (para evitar ruído em gráficos)
                # No Excel, podemos incluir todos os valores, mas a lógica de filtro do gráfico pode ser mantida.
                sheet.append([label, value])
                # A célula de valor está na coluna 2 da linha recém-adicionada
                sheet.cell(row=sheet.max_row, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE # Formato de moeda

        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name="relatorio_financeiro.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif format == 'pdf':
        # Importa apenas quando necessário
        from fpdf import FPDF
        import io

        # Configuração do PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size = 12)

        pdf.cell(200, 10, text = f"Relatório Financeiro de {start_date_str} a {end_date_str}", new_x="LMARGIN", new_y="NEXT", align = 'C')
        pdf.ln(10)

        # Detalhes das Transações
        if filtered_transactions:
            pdf.set_font("Arial", size = 10, style='B')
            pdf.cell(0, 10, text="Transações Detalhadas:", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Arial", size = 8)
            
            # Cabeçalho da tabela
            col_widths = [60, 20, 20, 20, 40, 30] # Ajuste as larguras conforme necessário
            pdf.cell(col_widths[0], 7, "Descrição", 1)
            pdf.cell(col_widths[1], 7, "Valor", 1)
            pdf.cell(col_widths[2], 7, "Data", 1)
            pdf.cell(col_widths[3], 7, "Tipo", 1)
            pdf.cell(col_widths[4], 7, "Categoria", 1)
            pdf.cell(col_widths[5], 7, "Conta", 1)
            pdf.ln()

            for t in filtered_transactions:
                category_name = t.category.name if t.category else "N/A"
                account_name = t.account.name if t.account else "N/A"
                
                # Truncar descrição para caber na célula e remover quebras de linha
                description_display = t.description.replace('\n', ' ').replace('\r', ' ')
                if len(description_display) > 35: 
                    description_display = description_display[:32] + "..."

                pdf.cell(col_widths[0], 7, description_display, 1)
                pdf.cell(col_widths[1], 7, f"R$ {t.amount:.2f}", 1)
                pdf.cell(col_widths[2], 7, t.date, 1)
                pdf.cell(col_widths[3], 7, "Receita" if t.type == "income" else "Despesa", 1)
                pdf.cell(col_widths[4], 7, category_name, 1)
                pdf.cell(col_widths[5], 7, account_name, 1)
                pdf.ln()
        else:
            pdf.cell(0, 10, text="Nenhuma transação encontrada para os filtros selecionados.", new_x="LMARGIN", new_y="NEXT")
        
        pdf.ln(10)

        # Resumo de Despesas por Categoria
        if report_data['expenses_by_category_chart']['labels']:
            pdf.set_font("Arial", size = 10, style='B')
            pdf.cell(0, 10, text="Despesas por Categoria:", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Arial", size = 8)
            for i, label in enumerate(report_data['expenses_by_category_chart']['labels']):
                value = report_data['expenses_by_category_chart']['values'][i]
                pdf.cell(0, 7, text=f"{label}: R$ {value:.2f}", new_x="LMARGIN", new_y="NEXT")
        
        pdf.ln(10)

        # Evolução do Patrimônio Líquido
        if report_data['net_worth_evolution_chart']['labels']:
            pdf.set_font("Arial", size = 10, style='B')
            pdf.cell(0, 10, text="Evolução do Patrimônio Líquido:", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Arial", size = 8)
            for i, label in enumerate(report_data['net_worth_evolution_chart']['labels']):
                value = report_data['net_worth_evolution_chart']['values'][i]
                # Filtrar valores: apenas se o valor for > 0.01 ou < 0.00 (para evitar ruído de pequenas flutuações)
                if value > 0.01 or value < 0.00:
                    pdf.cell(0, 7, text=f"{label}: R$ {value:.2f}", new_x="LMARGIN", new_y="NEXT")

        return send_file(io.BytesIO(pdf.output(dest='S')), download_name="relatorio_financeiro.pdf", as_attachment=True, mimetype='application/pdf')

    else:
        flash('Formato de exportação inválido.', 'danger')
        return redirect(url_for('report.reports_page'))

